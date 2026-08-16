from datetime import timedelta
from decimal import Decimal
from functools import wraps
from io import BytesIO
from urllib.parse import urlencode
import json

from allauth.account.models import EmailAddress
from django.conf import settings
from django.contrib import messages
from django.contrib.auth import get_user_model
from django.contrib.auth.views import redirect_to_login
from django.contrib.sessions.models import Session
from django.core.paginator import Paginator
from django.db.models import Count, Prefetch, Q, Sum
from django.db.models.functions import Coalesce, Lower, TruncDate, TruncMonth
from django.http import FileResponse, HttpResponse, HttpResponseForbidden, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.template.loader import render_to_string
from django.urls import reverse
from django.utils import timezone
from django.views.decorators.http import require_POST
from openpyxl import Workbook
from openpyxl.styles import Font

from .auth import (
    backoffice_required,
    emails_backoffice_autorises,
    emails_backoffice_env,
    invalider_cache_acces_backoffice,
)
from .partials import partial_kind as _partial_kind
from .partials import request_wants_ajax as _request_wants_ajax

# Compat historique (exports internes du monolithe)
_emails_backoffice_autorises = emails_backoffice_autorises

from ..models import (
    AccesBackoffice,
    AbonnementOrganisation,
    ChargePlateforme,
    MembreOrganisation,
    Organisation,
    PaiementAbonnement,
    PlanAbonnement,
    ProfilUtilisateur,
)

User = get_user_model()


def _debut_mois_local(dt=None):
    dt = timezone.localtime(dt or timezone.now())
    return dt.replace(day=1, hour=0, minute=0, second=0, microsecond=0)


def _add_months(dt, n):
    """Décale un datetime au 1er jour du mois + n mois (n peut être négatif)."""
    base = _debut_mois_local(dt)
    m0 = base.month - 1 + n
    year = base.year + m0 // 12
    month = m0 % 12 + 1
    return base.replace(year=year, month=month)


def _liste_mois(nb=12):
    """Liste des premiers jours de mois (local), du plus ancien au plus récent."""
    d = _debut_mois_local()
    mois = []
    for i in range(nb):
        mois.append(_add_months(d, -i))
    return list(reversed(mois))


def _liste_mois_entre(debut, fin):
    """Mois inclus (1er jour local) de debut à fin."""
    if debut is None or fin is None:
        return _liste_mois(12)
    start = _debut_mois_local(debut)
    end = _debut_mois_local(fin)
    if start > end:
        start, end = end, start
    mois = []
    cur = start
    for _ in range(120):
        mois.append(cur)
        if cur >= end:
            break
        cur = _add_months(cur, 1)
    return mois or [start]


def _parse_iso_date(value):
    """Parse YYYY-MM-DD → date, ou None."""
    raw = (value or '').strip()
    if not raw:
        return None
    try:
        parts = raw.split('-')
        if len(parts) != 3:
            return None
        y, m, d = int(parts[0]), int(parts[1]), int(parts[2])
        from datetime import date as date_cls
        return date_cls(y, m, d)
    except (TypeError, ValueError, OverflowError):
        return None


def _liste_acces_backoffice(viewer=None):
    """Liste unifiée des opérateurs (.env + DB) pour l’UI Outils."""
    from django.db.models.functions import Lower as LowerFn

    env_set = emails_backoffice_env()
    db_rows = {
        str(row.email).strip().lower(): row
        for row in AccesBackoffice.objects.all()
    }
    viewer_email = ''
    if viewer is not None and getattr(viewer, 'is_authenticated', False):
        viewer_email = (
            getattr(viewer, 'email', None) or viewer.get_username() or ''
        ).strip().lower()

    all_emails = sorted(set(env_set) | set(db_rows.keys()))

    users_by_email = {}
    if all_emails:
        UserModel = get_user_model()
        for u in UserModel.objects.annotate(email_l=LowerFn('email')).filter(
            email_l__in=all_emails
        ).only('id', 'email', 'first_name', 'last_name', 'username'):
            users_by_email[(u.email or '').strip().lower()] = u

    lignes = []
    for email in all_emails:
        row = db_rows.get(email)
        via_env = email in env_set
        via_db = bool(row and row.actif)
        if not via_env and not row:
            continue
        user = users_by_email.get(email)
        can_revoke = via_db and email != viewer_email
        can_delete = bool(row) and email != viewer_email
        can_reactivate = bool(row) and not row.actif and email != viewer_email
        if via_env and via_db:
            source_label = 'Env + backoffice'
        elif via_env:
            source_label = 'Fichier .env'
        elif via_db:
            source_label = 'Backoffice'
        else:
            source_label = 'Révoqué'
        lignes.append({
            'email': email,
            'via_env': via_env,
            'via_db': via_db,
            'actif': via_env or via_db,
            'db_id': row.pk if row else None,
            'note': (row.note if row else '') or '',
            'ajoute_le': row.cree_le if row else None,
            'user_id': user.pk if user else None,
            'user_label': (
                (user.get_full_name() or user.username) if user else ''
            ),
            'is_self': email == viewer_email,
            'can_revoke': can_revoke,
            'can_delete': can_delete,
            'can_reactivate': can_reactivate,
            'source_label': source_label,
        })
    lignes.sort(key=lambda x: (0 if x['actif'] else 1, x['email']))
    return {
        'lignes': lignes,
        'n_actifs': sum(1 for x in lignes if x['actif']),
        'viewer_email': viewer_email,
    }


def _local_day_start(d):
    """Début de journée locale (aware si USE_TZ) pour une date."""
    from datetime import datetime, time
    naive = datetime.combine(d, time.min)
    if settings.USE_TZ:
        return timezone.make_aware(naive, timezone.get_current_timezone())
    return naive


def _local_day_end(d):
    from datetime import datetime, time
    naive = datetime.combine(d, time(23, 59, 59, 999999))
    if settings.USE_TZ:
        return timezone.make_aware(naive, timezone.get_current_timezone())
    return naive


def _iso_date(d):
    if d is None:
        return ''
    return d.isoformat() if hasattr(d, 'isoformat') else str(d)


def _format_date_fr(d):
    if d is None:
        return ''
    mois_fr = (
        'janvier', 'février', 'mars', 'avril', 'mai', 'juin',
        'juillet', 'août', 'septembre', 'octobre', 'novembre', 'décembre',
    )
    return f'{d.day} {mois_fr[d.month - 1]} {d.year}'


# Présélections du filtre temporel backoffice
PERIODE_PRESETS = (
    ('all', 'Tout'),
    ('today', "Aujourd'hui"),
    ('7j', '7 jours'),
    ('30j', '30 jours'),
    ('month', 'Ce mois'),
    ('quarter', 'Trimestre'),
    ('year', 'Année'),
    ('12m', '12 mois'),
)
PERIODE_KEYS = {k for k, _ in PERIODE_PRESETS}


def _debut_historique_plateforme(local_now):
    first_user = (
        User.objects.order_by('date_joined')
        .values_list('date_joined', flat=True)
        .first()
    )
    first_org = (
        Organisation.objects.order_by('cree_le')
        .values_list('cree_le', flat=True)
        .first()
    )
    candidates = [d for d in (first_user, first_org) if d]
    if candidates:
        debut = min(candidates)
        # Limite graphiques : 5 ans
        min_charts = _add_months(local_now, -59)
        if debut < min_charts:
            debut = min_charts
        return debut
    return _add_months(local_now, -11)


def _trouver_preset_pour_dates(du_date, au_date, today, local_now):
    """
    Si le calendrier coïncide avec un preset, renvoyer sa clé (pour le highlight UI).
    Sinon 'custom'.
    """
    if not du_date or not au_date:
        return 'custom'
    if du_date == au_date == today:
        return 'today'
    if du_date == today - timedelta(days=6) and au_date == today:
        return '7j'
    if du_date == today - timedelta(days=29) and au_date == today:
        return '30j'
    if du_date == today.replace(day=1) and au_date == today:
        return 'month'
    q = (today.month - 1) // 3 * 3 + 1
    if du_date == today.replace(month=q, day=1) and au_date == today:
        return 'quarter'
    if du_date == today.replace(month=1, day=1) and au_date == today:
        return 'year'
    debut_12m_dt = _add_months(local_now, -11)
    debut_12m = (
        timezone.localtime(debut_12m_dt).date()
        if timezone.is_aware(debut_12m_dt)
        else debut_12m_dt.date()
    )
    if du_date == debut_12m and au_date == today:
        return '12m'
    return 'custom'


def _periode_depuis_request(request):
    """
    Bornes temporelles pour KPI / graphiques / liste.
    Priorité : dates du/au (calendrier) > preset periode.
    Défaut : toutes les données (all).
    Retourne un dict unifié.
    """
    now = timezone.now()
    local_now = timezone.localtime(now)
    today = local_now.date()
    du_raw = (request.GET.get('du') or '').strip()
    au_raw = (request.GET.get('au') or '').strip()
    du_date = _parse_iso_date(du_raw)
    au_date = _parse_iso_date(au_raw)

    # —— Calendrier personnalisé ——
    if du_date or au_date:
        if du_date and au_date and du_date > au_date:
            du_date, au_date = au_date, du_date
        if du_date and not au_date:
            au_date = du_date
        if au_date and not du_date:
            du_date = au_date
        debut = _local_day_start(du_date)
        fin = _local_day_end(au_date)
        if du_date == au_date:
            label = f'Jour du {_format_date_fr(du_date)}'
            mode = 'day'
        else:
            label = f'Du {_format_date_fr(du_date)} au {_format_date_fr(au_date)}'
            mode = 'range'
        # Align highlight avec un preset si les dates coïncident
        key = _trouver_preset_pour_dates(du_date, au_date, today, local_now)
        if key == 'today':
            label = "Aujourd'hui"
        elif key == '7j':
            label = '7 derniers jours'
        elif key == '30j':
            label = '30 derniers jours'
        elif key == 'month':
            label = 'Ce mois'
        elif key == 'quarter':
            label = 'Ce trimestre'
        elif key == 'year':
            label = 'Cette année'
        elif key == '12m':
            label = '12 derniers mois'
        return {
            'key': key,
            'debut': debut,
            'fin': fin,
            'label': label,
            'du': _iso_date(du_date),
            'au': _iso_date(au_date),
            'mode': mode,
            'filtre_actif': True,
            'charts_debut': debut,
            'charts_fin': fin,
        }

    raw = (request.GET.get('periode') or 'all').strip().lower()
    key = raw if raw in PERIODE_KEYS else 'all'
    fin = now

    if key == 'today':
        debut = _local_day_start(today)
        fin = _local_day_end(today)
        label = "Aujourd'hui"
        du, au = today, today
        mode = 'day'
    elif key == '7j':
        d0 = today - timedelta(days=6)
        debut = _local_day_start(d0)
        fin = _local_day_end(today)
        label = '7 derniers jours'
        du, au = d0, today
        mode = 'range'
    elif key == '30j':
        d0 = today - timedelta(days=29)
        debut = _local_day_start(d0)
        fin = _local_day_end(today)
        label = '30 derniers jours'
        du, au = d0, today
        mode = 'range'
    elif key == 'month':
        d0 = today.replace(day=1)
        debut = _local_day_start(d0)
        fin = _local_day_end(today)
        label = 'Ce mois'
        du, au = d0, today
        mode = 'range'
    elif key == 'quarter':
        q = (today.month - 1) // 3 * 3 + 1
        d0 = today.replace(month=q, day=1)
        debut = _local_day_start(d0)
        fin = _local_day_end(today)
        label = 'Ce trimestre'
        du, au = d0, today
        mode = 'range'
    elif key == 'year':
        d0 = today.replace(month=1, day=1)
        debut = _local_day_start(d0)
        fin = _local_day_end(today)
        label = 'Cette année'
        du, au = d0, today
        mode = 'range'
    elif key == '12m':
        debut = _add_months(local_now, -11)
        fin = _local_day_end(today)
        label = '12 derniers mois'
        du = timezone.localtime(debut).date()
        au = today
        mode = 'range'
    else:
        # all — pas de filtre liste/KPI flux strict ; graphiques sur l’historique
        return {
            'key': 'all',
            'debut': None,
            'fin': None,
            'label': 'Toutes les données',
            'du': '',
            'au': '',
            'mode': 'range',
            'filtre_actif': False,
            'charts_debut': _debut_historique_plateforme(local_now),
            'charts_fin': now,
        }

    return {
        'key': key,
        'debut': debut,
        'fin': fin,
        'label': label,
        'du': _iso_date(du),
        'au': _iso_date(au),
        'mode': mode,
        'filtre_actif': True,
        'charts_debut': debut,
        'charts_fin': fin,
    }


def _params_liste_utilisateurs(filtre_actif, periode_debut, periode_fin, user_scope):
    """
    Règles uniques liste / export :
    - sessions → en ligne maintenant (ignore le calendrier)
    - sinon période si active (scopes inscrits/connectes/activité)
    """
    if user_scope == 'sessions':
        return None, None, 'sessions'
    if filtre_actif and periode_debut is not None and periode_fin is not None:
        scope = user_scope if user_scope in ('inscrits', 'connectes') else ''
        return periode_debut, periode_fin, scope
    return None, None, ''


def _repartition_geo_utilisateurs(debut=None, fin=None, limite_pays=12, limite_villes=15):
    """
    Répartition pays / ville depuis ProfilUtilisateur.
    Si debut/fin fournis : limité aux comptes actifs sur la période (inscrits ou login).
    """
    users = User.objects.all()
    if debut is not None and fin is not None:
        users = users.filter(_q_users_activite_periode(debut, fin))
    user_ids = list(users.values_list('pk', flat=True))
    total = len(user_ids)
    if not user_ids:
        return {
            'total': 0,
            'localises': 0,
            'sans_localisation': 0,
            'pct_localises': 0,
            'pays': {'labels': [], 'values': []},
            'villes': [],
        }

    profils = ProfilUtilisateur.objects.filter(utilisateur_id__in=user_ids)
    localises_ids = set(
        profils.exclude(pays='').values_list('utilisateur_id', flat=True)
    )
    localises = len(localises_ids)
    sans = max(0, total - localises)

    pays_rows = list(
        profils.exclude(pays='')
        .values('pays')
        .annotate(n=Count('id'))
        .order_by('-n', 'pays')[:limite_pays]
    )
    ville_rows = list(
        profils.exclude(pays='').exclude(ville='')
        .values('pays', 'ville')
        .annotate(n=Count('id'))
        .order_by('-n', 'pays', 'ville')[:limite_villes]
    )

    base_pays = localises or 1
    villes = []
    for row in ville_rows:
        n = int(row['n'] or 0)
        villes.append({
            'pays': row['pays'],
            'ville': row['ville'],
            'n': n,
            'label': f"{row['ville']} · {row['pays']}",
            'pct': round(100 * n / base_pays) if base_pays else 0,
        })

    return {
        'total': total,
        'localises': localises,
        'sans_localisation': sans,
        'pct_localises': round(100 * localises / total) if total else 0,
        'pays': {
            'labels': [r['pays'] for r in pays_rows],
            'values': [int(r['n'] or 0) for r in pays_rows],
        },
        'villes': villes,
    }



def _as_local_date(value):
    """Normalise date ou datetime → date locale."""
    if value is None:
        return None
    if hasattr(value, 'hour'):
        if timezone.is_aware(value):
            return timezone.localtime(value).date()
        if settings.USE_TZ:
            return timezone.make_aware(
                value, timezone.get_current_timezone()
            ).astimezone(timezone.get_current_timezone()).date()
        return value.date()
    return value


def _label_mois(d):
    d = _as_local_date(d) or d
    mois_fr = (
        'janv.', 'févr.', 'mars', 'avr.', 'mai', 'juin',
        'juil.', 'août', 'sept.', 'oct.', 'nov.', 'déc.',
    )
    return f'{mois_fr[d.month - 1]} {d.year % 100:02d}'


def _label_jour(d):
    d = _as_local_date(d) or d
    return f'{d.day:02d}/{d.month:02d}'


def _liste_jours_entre(debut, fin):
    """Dates civiles (locale) de debut à fin inclus, max 400 jours."""
    d0 = _as_local_date(debut)
    d1 = _as_local_date(fin)
    if d0 is None or d1 is None:
        return []
    if d0 > d1:
        d0, d1 = d1, d0
    days = []
    cur = d0
    for _ in range(400):
        days.append(cur)
        if cur >= d1:
            break
        cur = cur + timedelta(days=1)
    return days


def _granularite_charts(debut, fin):
    """Jour si fenêtre ≤ 45 j (today / 7j / 30j / calendrier court), sinon mois."""
    d0 = _as_local_date(debut)
    d1 = _as_local_date(fin)
    if d0 is None or d1 is None:
        return 'month'
    span = (d1 - d0).days + 1
    return 'day' if span <= 45 else 'month'


def _bucket_key(value, granularite):
    if value is None:
        return None
    if granularite == 'day':
        return _as_local_date(value)
    local = value
    if hasattr(value, 'hour'):
        local = timezone.localtime(value) if timezone.is_aware(value) else value
        return (local.year, local.month)
    d = _as_local_date(value)
    return (d.year, d.month) if d else None


def _point_key(point, granularite):
    if granularite == 'day':
        return _as_local_date(point)
    local = point
    if hasattr(point, 'hour'):
        local = timezone.localtime(point) if timezone.is_aware(point) else point
    else:
        local = point
    return (local.year, local.month)


def _compter_par_buckets(queryset, date_field, points, granularite, debut, fin):
    """Compte groupé par jour ou mois, borné strictement à [debut, fin]."""
    if not points or debut is None or fin is None:
        return [0] * len(points or [])
    trunc = TruncDate(date_field) if granularite == 'day' else TruncMonth(date_field)
    rows = (
        queryset.filter(**{
            f'{date_field}__gte': debut,
            f'{date_field}__lte': fin,
        })
        .annotate(bucket=trunc)
        .values('bucket')
        .annotate(n=Count('id'))
    )
    by_key = {}
    for row in rows:
        key = _bucket_key(row['bucket'], granularite)
        if key is not None:
            by_key[key] = int(row['n'] or 0)
    return [by_key.get(_point_key(p, granularite), 0) for p in points]


def _sommer_paiements_par_buckets(points, granularite, debut, fin):
    """Somme des paiements réussis (date = paye_le ou cree_le) par bucket."""
    if not points or debut is None or fin is None:
        return [0.0] * len(points or [])
    qs = (
        PaiementAbonnement.objects
        .filter(statut=PaiementAbonnement.STATUT_REUSSI)
        .annotate(dt_pay=Coalesce('paye_le', 'cree_le'))
        .filter(dt_pay__gte=debut, dt_pay__lte=fin)
    )
    trunc = TruncDate('dt_pay') if granularite == 'day' else TruncMonth('dt_pay')
    rows = (
        qs.annotate(bucket=trunc)
        .values('bucket')
        .annotate(total=Sum('montant'))
    )
    by_key = {}
    for row in rows:
        key = _bucket_key(row['bucket'], granularite)
        if key is not None:
            by_key[key] = float(row['total'] or 0)
    return [by_key.get(_point_key(p, granularite), 0.0) for p in points]


def _dates_periode_calendaire(debut, fin):
    """Bornes date (local) pour filtrer ChargePlateforme.date_charge."""
    if debut is None or fin is None:
        return None, None
    return timezone.localtime(debut).date(), timezone.localtime(fin).date()


def _charges_periode(debut=None, fin=None):
    qs = ChargePlateforme.objects.all()
    d0, d1 = _dates_periode_calendaire(debut, fin)
    if d0 is not None and d1 is not None:
        qs = qs.filter(date_charge__gte=d0, date_charge__lte=d1)
    return qs


def _sommer_charges_par_buckets(points, granularite, debut, fin):
    """Somme des charges plateforme par bucket (date_charge)."""
    if not points or debut is None or fin is None:
        return [0.0] * len(points or [])
    d0, d1 = _dates_periode_calendaire(debut, fin)
    if d0 is None or d1 is None:
        return [0.0] * len(points)
    qs = ChargePlateforme.objects.filter(date_charge__gte=d0, date_charge__lte=d1)
    trunc = TruncDate('date_charge') if granularite == 'day' else TruncMonth('date_charge')
    rows = (
        qs.annotate(bucket=trunc)
        .values('bucket')
        .annotate(total=Sum('montant'))
    )
    by_key = {}
    for row in rows:
        key = _bucket_key(row['bucket'], granularite)
        if key is not None:
            by_key[key] = float(row['total'] or 0)
    return [by_key.get(_point_key(p, granularite), 0.0) for p in points]


def _stats_charges(qs):
    agg = qs.aggregate(total=Sum('montant'), n=Count('id'))
    total = agg.get('total') or Decimal('0')
    return {
        'n': agg.get('n') or 0,
        'total': total,
        'total_formate': _format_montant(total),
    }


def _queryset_charges_filtres(charge_q='', charge_categorie='', debut=None, fin=None):
    qs = _charges_periode(debut, fin)
    cat_codes = {c for c, _ in ChargePlateforme.CATEGORIE_CHOICES}
    if charge_categorie in cat_codes:
        qs = qs.filter(categorie=charge_categorie)
    if charge_q:
        qs = qs.filter(
            Q(libelle__icontains=charge_q) | Q(notes__icontains=charge_q)
        )
    return qs.order_by('-date_charge', '-cree_le', '-pk')


def _lignes_charges(qs):
    cat_labels = dict(ChargePlateforme.CATEGORIE_CHOICES)
    lignes = []
    for ch in qs:
        lignes.append({
            'id': ch.pk,
            'date_label': ch.date_charge.strftime('%d/%m/%Y'),
            'date_iso': ch.date_charge.isoformat(),
            'libelle': ch.libelle,
            'categorie': ch.categorie,
            'categorie_label': cat_labels.get(ch.categorie, ch.categorie),
            'montant': _format_montant(ch.montant, ch.devise),
            'montant_raw': str(int(ch.montant) if ch.montant == int(ch.montant) else ch.montant),
            'notes': (ch.notes or '').strip(),
        })
    return lignes


def _parse_charge_payload(post):
    """Valide les champs charge (create / update)."""
    raw_date = (post.get('date_charge') or '').strip()
    raw_montant = (post.get('montant') or '').strip().replace(' ', '').replace(',', '.')
    categorie = (post.get('categorie') or ChargePlateforme.CAT_AUTRE).strip()
    libelle = (post.get('libelle') or '').strip()
    notes = (post.get('notes') or '').strip()

    date_charge = _parse_iso_date(raw_date)
    if date_charge is None:
        raise ValueError('Date invalide.')
    if not libelle:
        raise ValueError('Libellé obligatoire.')
    try:
        montant = Decimal(raw_montant)
    except Exception as exc:
        raise ValueError('Montant invalide.') from exc
    if montant <= 0:
        raise ValueError('Le montant doit être supérieur à zéro.')

    cat_codes = {c for c, _ in ChargePlateforme.CATEGORIE_CHOICES}
    if categorie not in cat_codes:
        categorie = ChargePlateforme.CAT_AUTRE

    return {
        'date_charge': date_charge,
        'montant': montant,
        'categorie': categorie,
        'libelle': libelle[:200],
        'notes': notes,
    }


def _base_revenus_avant(debut):
    """Cumul paiements réussis strictement avant debut (même règle de date)."""
    if debut is None:
        return Decimal('0')
    total = (
        PaiementAbonnement.objects
        .filter(statut=PaiementAbonnement.STATUT_REUSSI)
        .annotate(dt_pay=Coalesce('paye_le', 'cree_le'))
        .filter(dt_pay__lt=debut)
        .aggregate(total=Sum('montant'))
        .get('total')
    )
    return total or Decimal('0')


def _graphiques_plateforme(debut=None, fin=None, nb_mois=12):
    """
    Séries graphiques backoffice pour la fenêtre [debut, fin].
    Granularité auto : jour (≤ 45 j) ou mois (longues plages / historique).
    """
    now = timezone.now()
    if debut is None or fin is None:
        points = _liste_mois(nb_mois)
        debut = points[0] if points else _add_months(now, -(nb_mois - 1))
        fin = now
        granularite = 'month'
    else:
        granularite = _granularite_charts(debut, fin)
        if granularite == 'day':
            points = _liste_jours_entre(debut, fin)
        else:
            points = _liste_mois_entre(debut, fin)

    if granularite == 'day':
        labels = [_label_jour(d) for d in points]
        unit_label = 'jour'
    else:
        labels = [_label_mois(d) for d in points]
        unit_label = 'mois'

    inscrits = _compter_par_buckets(
        User.objects.all(), 'date_joined', points, granularite, debut, fin,
    )
    orgs = _compter_par_buckets(
        Organisation.objects.all(), 'cree_le', points, granularite, debut, fin,
    )
    connectes = _compter_par_buckets(
        User.objects.exclude(last_login__isnull=True),
        'last_login',
        points,
        granularite,
        debut,
        fin,
    )
    revenus = _sommer_paiements_par_buckets(points, granularite, debut, fin)
    charges = _sommer_charges_par_buckets(points, granularite, debut, fin)
    resultat_net = [
        float(r or 0) - float(c or 0) for r, c in zip(revenus, charges)
    ]

    base_inscrits = User.objects.filter(date_joined__lt=debut).count()
    cumul = base_inscrits
    inscrits_cumul = []
    for n in inscrits:
        cumul += n
        inscrits_cumul.append(cumul)

    base_rev = float(_base_revenus_avant(debut))
    run_rev = base_rev
    revenus_cumul = []
    for v in revenus:
        run_rev += float(v or 0)
        revenus_cumul.append(run_rev)

    total_revenus = sum(float(v or 0) for v in revenus)
    total_charges = sum(float(v or 0) for v in charges)
    total_resultat = total_revenus - total_charges
    total_inscr = sum(inscrits)
    total_orgs = sum(orgs)
    total_conn = sum(connectes)

    return {
        'labels': labels,
        'granularite': granularite,
        'unit_label': unit_label,
        'inscrits_mois': inscrits,
        'inscrits_cumul': inscrits_cumul,
        'orgs_mois': orgs,
        'revenus_mois': revenus,
        'revenus_cumul': revenus_cumul,
        'charges_mois': charges,
        'resultat_net_mois': resultat_net,
        'connectes_mois': connectes,
        'totaux': {
            'revenus': total_revenus,
            'revenus_formate': _format_montant(Decimal(str(total_revenus))),
            'charges': total_charges,
            'charges_formate': _format_montant(Decimal(str(total_charges))),
            'resultat_net': total_resultat,
            'resultat_net_formate': _format_montant(Decimal(str(total_resultat))),
            'inscriptions': total_inscr,
            'orgs': total_orgs,
            'connexions': total_conn,
        },
    }


def _format_montant(valeur, devise='XOF'):
    montant = valeur or Decimal('0')
    try:
        entier = int(montant)
    except (TypeError, ValueError):
        entier = 0
    texte = f'{entier:,}'.replace(',', ' ')
    return f'{texte} {devise}'


def _jours_relatifs(dt_cible, maintenant=None):
    """Nombre de jours calendaires (local) de maintenant → dt_cible (peut être négatif)."""
    if dt_cible is None:
        return None
    maintenant = maintenant or timezone.now()
    d_cible = timezone.localtime(dt_cible).date()
    d_now = timezone.localtime(maintenant).date()
    return (d_cible - d_now).days


def _libelle_jours_relatifs(jours):
    if jours is None:
        return ''
    if jours == 0:
        return "aujourd'hui"
    if jours == 1:
        return 'demain'
    if jours > 1:
        return f'dans {jours} j'
    if jours == -1:
        return 'hier'
    return f'il y a {-jours} j'


def _format_dt_fr(dt):
    if dt is None:
        return '—'
    local = timezone.localtime(dt)
    return _format_date_fr(local)


def _contact_organisation(org):
    """E-mail / nom du propriétaire (fallback admin puis e-mail org)."""
    if org is None:
        return {'email': '', 'nom': '', 'user_id': None}
    prefetched = getattr(org, '_prefetched_objects_cache', {})
    if 'membres' in prefetched:
        membres = [m for m in org.membres.all() if m.actif]
    else:
        membres = list(
            org.membres.filter(actif=True)
            .select_related('utilisateur')
            .order_by('id')[:8]
        )
    role_rank = {
        MembreOrganisation.ROLE_PROPRIETAIRE: 0,
        MembreOrganisation.ROLE_ADMIN: 1,
        MembreOrganisation.ROLE_MEMBRE: 2,
    }
    membres.sort(key=lambda m: (role_rank.get(m.role, 9), m.pk or 0))
    prefer = membres[0] if membres else None
    if prefer is None:
        return {
            'email': (org.email or '').strip(),
            'nom': '',
            'user_id': None,
        }
    user = prefer.utilisateur
    return {
        'email': (prefer.get_email() or '').strip(),
        'nom': (prefer.get_nom_affichage() or '').strip(),
        'user_id': user.pk if user else None,
    }


def _users_href(params):
    """Lien vers l’onglet Utilisateurs avec query string."""
    qs = {k: v for k, v in (params or {}).items() if v not in (None, '', False)}
    if not qs:
        return '#utilisateurs'
    return f'?{urlencode(qs)}#utilisateurs'


def _alertes_abonnements(maintenant=None, horizon_jours=7, echecs_jours=7, limite=12):
    """
    Alertes actionnables pour la Vue d'ensemble :
    - échéances proche (essai / actif dans horizon_jours)
    - abonnements en retard
    - paiements échoués récents
    """
    maintenant = maintenant or timezone.now()
    horizon = maintenant + timedelta(days=max(1, int(horizon_jours)))
    debut_echecs = maintenant - timedelta(days=max(1, int(echecs_jours)))
    limite = max(1, int(limite))

    membres_qs = MembreOrganisation.objects.filter(actif=True).select_related('utilisateur')

    abo_base = AbonnementOrganisation.objects.select_related(
        'organisation', 'plan',
    ).prefetch_related(
        Prefetch('organisation__membres', queryset=membres_qs),
    )

    # 1) Expire bientôt — uniquement si statut effectif encore essai/actif
    expire_qs = abo_base.filter(
        Q(
            statut=AbonnementOrganisation.STATUT_ESSAI,
            essai_fin__isnull=False,
            essai_fin__gte=maintenant,
            essai_fin__lte=horizon,
        )
        | Q(
            statut=AbonnementOrganisation.STATUT_ACTIF,
            periode_fin__isnull=False,
            periode_fin__gte=maintenant,
            periode_fin__lte=horizon,
        )
    ).annotate(
        _fin_proche=Coalesce('essai_fin', 'periode_fin'),
    ).order_by('_fin_proche')
    expire_items = []
    n_expire = 0
    for abo in expire_qs:
        eff = _statut_effectif(abo, maintenant)
        if eff not in (
            AbonnementOrganisation.STATUT_ESSAI,
            AbonnementOrganisation.STATUT_ACTIF,
        ):
            continue
        n_expire += 1
        if len(expire_items) >= limite:
            continue
        fin = (
            abo.essai_fin
            if eff == AbonnementOrganisation.STATUT_ESSAI
            else abo.periode_fin
        )
        jours = _jours_relatifs(fin, maintenant)
        contact = _contact_organisation(abo.organisation)
        search_q = contact['email'] or abo.organisation.nom or ''
        expire_items.append({
            'type': 'expire_bientot',
            'org_nom': abo.organisation.nom or abo.organisation.slug,
            'org_id': abo.organisation_id,
            'plan': abo.plan.nom if abo.plan_id else '—',
            'statut': eff,
            'statut_label': _statut_label(eff),
            'echeance': _format_dt_fr(fin),
            'jours': jours,
            'jours_label': _libelle_jours_relatifs(jours),
            'contact_email': contact['email'],
            'contact_nom': contact['nom'],
            'href_users': _users_href({'q': search_q, 'statut': abo.statut}),
            'href_org': reverse('backoffice_org_detail', args=[abo.organisation_id]),
            'resume': (
                'essai se termine'
                if eff == AbonnementOrganisation.STATUT_ESSAI
                else 'période se termine'
            ),
            'severity': 'warning',
        })

    # 2) En retard effectif (base en_retard ou actif hors période, encore dans la grâce)
    retard_candidats = abo_base.filter(
        Q(statut=AbonnementOrganisation.STATUT_EN_RETARD)
        | Q(
            statut=AbonnementOrganisation.STATUT_ACTIF,
            periode_fin__isnull=False,
            periode_fin__lt=maintenant,
        )
    ).order_by('periode_fin')
    retard_items = []
    n_retard = 0
    for abo in retard_candidats:
        if _statut_effectif(abo, maintenant) != AbonnementOrganisation.STATUT_EN_RETARD:
            continue
        n_retard += 1
        if len(retard_items) >= limite:
            continue
        fin = abo.periode_fin
        fin_grace = abo.fin_grace_retard()
        jours_depuis = _jours_relatifs(fin, maintenant)
        jours_grace = _jours_relatifs(fin_grace, maintenant)
        contact = _contact_organisation(abo.organisation)
        search_q = contact['email'] or abo.organisation.nom or ''
        if jours_depuis is not None and jours_depuis < 0:
            jours_label = f'depuis {-jours_depuis} j'
        else:
            jours_label = _libelle_jours_relatifs(jours_depuis)
        retard_items.append({
            'type': 'en_retard',
            'org_nom': abo.organisation.nom or abo.organisation.slug,
            'org_id': abo.organisation_id,
            'plan': abo.plan.nom if abo.plan_id else '—',
            'statut': AbonnementOrganisation.STATUT_EN_RETARD,
            'statut_label': _statut_label(AbonnementOrganisation.STATUT_EN_RETARD),
            'echeance': _format_dt_fr(fin),
            'grace': _format_dt_fr(fin_grace),
            'jours': jours_depuis,
            'jours_label': jours_label,
            'grace_label': (
                f'grâce {_libelle_jours_relatifs(jours_grace)}'
                if jours_grace is not None
                else 'grâce non définie'
            ),
            'contact_email': contact['email'],
            'contact_nom': contact['nom'],
            'href_users': _users_href({
                'q': search_q,
                'statut': AbonnementOrganisation.STATUT_EN_RETARD,
            }),
            'href_org': reverse('backoffice_org_detail', args=[abo.organisation_id]),
            'resume': 'paiement en retard',
            'severity': 'danger',
        })

    # 3) Paiements échoués récents (1 ligne / org — le plus récent)
    echecs_qs = (
        PaiementAbonnement.objects.filter(
            statut=PaiementAbonnement.STATUT_ECHEC,
            cree_le__gte=debut_echecs,
        )
        .select_related('organisation', 'abonnement', 'abonnement__plan')
        .order_by('-cree_le')
    )
    n_echecs = echecs_qs.count()
    echec_org_ids = []
    for oid in echecs_qs.values_list('organisation_id', flat=True):
        if oid not in echec_org_ids:
            echec_org_ids.append(oid)
        if len(echec_org_ids) >= limite:
            break
    orgs_membres = {
        o.pk: o
        for o in Organisation.objects.filter(pk__in=echec_org_ids).prefetch_related(
            Prefetch('membres', queryset=membres_qs),
        )
    }
    echec_items = []
    vues_org = set()
    for p in echecs_qs:
        if p.organisation_id in vues_org:
            continue
        vues_org.add(p.organisation_id)
        org = orgs_membres.get(p.organisation_id) or p.organisation
        contact = _contact_organisation(org)
        search_q = contact['email'] or org.nom or ''
        abo = p.abonnement
        echec_items.append({
            'type': 'echec_paiement',
            'org_nom': org.nom or org.slug,
            'org_id': org.pk,
            'plan': abo.plan.nom if abo and abo.plan_id else '—',
            'statut': abo.statut if abo else '',
            'statut_label': abo.get_statut_display() if abo else '—',
            'montant': _format_montant(p.montant, p.devise or 'XOF'),
            'methode': p.methode or '—',
            'reference': p.reference_externe or '',
            'echeance': _format_dt_fr(p.cree_le),
            'jours': _jours_relatifs(p.cree_le, maintenant),
            'jours_label': _libelle_jours_relatifs(_jours_relatifs(p.cree_le, maintenant)),
            'contact_email': contact['email'],
            'contact_nom': contact['nom'],
            'href_users': _users_href({'q': search_q}),
            'href_org': reverse('backoffice_org_detail', args=[org.pk]),
            'resume': 'paiement échoué',
            'severity': 'danger',
        })
        if len(echec_items) >= limite:
            break

    return {
        'total': n_expire + n_retard + n_echecs,
        'horizon_jours': horizon_jours,
        'echecs_jours': echecs_jours,
        'expire_bientot': {
            'n': n_expire,
            'items': expire_items,
        },
        'en_retard': {
            'n': n_retard,
            'items': retard_items,
            'href_all': _users_href({'statut': AbonnementOrganisation.STATUT_EN_RETARD}),
        },
        'echecs': {
            'n': n_echecs,
            'items': echec_items,
        },
    }


def _ids_users_email_verifie():
    return set(
        EmailAddress.objects.filter(verified=True).values_list('user_id', flat=True)
    )


def _emails_verifies():
    return set(
        EmailAddress.objects.filter(verified=True)
        .annotate(email_l=Lower('email'))
        .values_list('email_l', flat=True)
    )


def _date_relative(dt, maintenant):
    if not dt:
        return 'jamais'
    delta = maintenant - dt
    secondes = int(delta.total_seconds())
    if secondes < 0:
        return 'à venir'
    if secondes < 60:
        return "à l'instant"
    if secondes < 3600:
        return f'il y a {secondes // 60} min'
    if secondes < 86400:
        h = secondes // 3600
        return f'il y a {h} h'
    jours = secondes // 86400
    if jours < 30:
        return f'il y a {jours} j'
    mois = jours // 30
    if mois < 12:
        return f'il y a {mois} mois'
    ans = jours // 365
    return f'il y a {ans} an' + ('s' if ans > 1 else '')


def _membre_principal(user):
    membres = list(user.membres_organisations.all())
    if not membres:
        return None
    ordre = {
        MembreOrganisation.ROLE_PROPRIETAIRE: 0,
        MembreOrganisation.ROLE_ADMIN: 1,
        MembreOrganisation.ROLE_MEMBRE: 2,
    }
    membres.sort(key=lambda m: (0 if m.actif else 1, ordre.get(m.role, 9), -m.pk))
    return membres[0]


def _abonnement_of(org):
    if org is None:
        return None
    try:
        return org.abonnement
    except AbonnementOrganisation.DoesNotExist:
        return None


def _statut_label(code):
    for c, label in AbonnementOrganisation.STATUT_CHOICES:
        if c == code:
            return label
    return code or '—'


# Courtes explications pour les chips « Répartition abonnements »
STATUT_ABO_AIDE = {
    AbonnementOrganisation.STATUT_PRELAUNCH: (
        'Avant le lancement officiel : accès ouvert, sans essai daté.'
    ),
    AbonnementOrganisation.STATUT_ESSAI: (
        'Période d’essai en cours : accès Pro jusqu’à la fin de l’essai.'
    ),
    AbonnementOrganisation.STATUT_ACTIF: (
        'Abonnement payant en cours de validité.'
    ),
    AbonnementOrganisation.STATUT_EN_RETARD: (
        'Période payante dépassée : encore en fenêtre de grâce.'
    ),
    AbonnementOrganisation.STATUT_EXPIRE: (
        'Plus d’accès : essai ou période (et grâce) terminés.'
    ),
    AbonnementOrganisation.STATUT_ANNULE: (
        'Abonnement annulé manuellement : sans accès produit.'
    ),
}

# Phrase exacte à taper pour confirmer le lancement global (anti-erreur).
LANCEMENT_CONFIRM_PHRASE = 'LANCER XALISS'
PROLONGER_TOUS_CONFIRM_PHRASE = 'PROLONGER TOUS'


def _statut_effectif(abo, maintenant=None):
    """
    Statut « métier » d’après les dates, sans écrire en base.
    Même règles que synchroniser_statut (essai dépassé, période, grâce).
    """
    if not abo:
        return ''
    maintenant = maintenant or timezone.now()
    s = abo.statut
    if s == AbonnementOrganisation.STATUT_ESSAI:
        if abo.essai_fin is not None and maintenant > abo.essai_fin:
            return AbonnementOrganisation.STATUT_EXPIRE
        return s
    if s == AbonnementOrganisation.STATUT_ACTIF:
        if abo.periode_fin is not None and maintenant > abo.periode_fin:
            s = AbonnementOrganisation.STATUT_EN_RETARD
        else:
            return s
    if s == AbonnementOrganisation.STATUT_EN_RETARD:
        fin_grace = abo.fin_grace_retard()
        if fin_grace is None or maintenant > fin_grace:
            return AbonnementOrganisation.STATUT_EXPIRE
        return AbonnementOrganisation.STATUT_EN_RETARD
    return s


def _q_statut_abonnement_effectif(code, maintenant=None, prefix=''):
    """
    Filtre ORM équivalent à _statut_effectif, pour lier KPI / liste / chips.
    prefix: '' sur AbonnementOrganisation, ou
    'membres_organisations__organisation__abonnement__' depuis User.
    """
    maintenant = maintenant or timezone.now()
    code = (code or '').strip()
    if not code:
        return Q()
    p = prefix or ''

    def f(field):
        return f'{p}{field}'

    # Seuil fin de grâce : periode_fin + N jours >= now  ⇔  periode_fin >= now - N jours
    grace = AbonnementOrganisation.duree_grace_retard()
    seuil_grace = maintenant - grace

    if code == AbonnementOrganisation.STATUT_PRELAUNCH:
        return Q(**{f('statut'): AbonnementOrganisation.STATUT_PRELAUNCH})

    if code == AbonnementOrganisation.STATUT_ANNULE:
        return Q(**{f('statut'): AbonnementOrganisation.STATUT_ANNULE})

    if code == AbonnementOrganisation.STATUT_ESSAI:
        return (
            Q(**{f('statut'): AbonnementOrganisation.STATUT_ESSAI})
            & (
                Q(**{f('essai_fin') + '__isnull': True})
                | Q(**{f('essai_fin') + '__gte': maintenant})
            )
        )

    if code == AbonnementOrganisation.STATUT_ACTIF:
        return (
            Q(**{f('statut'): AbonnementOrganisation.STATUT_ACTIF})
            & (
                Q(**{f('periode_fin') + '__isnull': True})
                | Q(**{f('periode_fin') + '__gte': maintenant})
            )
        )

    if code == AbonnementOrganisation.STATUT_EN_RETARD:
        # (stock en_retard encore dans la grâce) OU (stock actif, période passée, grâce OK)
        return (
            (
                Q(**{f('statut'): AbonnementOrganisation.STATUT_EN_RETARD})
                & Q(**{f('periode_fin') + '__isnull': False})
                & Q(**{f('periode_fin') + '__gte': seuil_grace})
            )
            | (
                Q(**{f('statut'): AbonnementOrganisation.STATUT_ACTIF})
                & Q(**{f('periode_fin') + '__isnull': False})
                & Q(**{f('periode_fin') + '__lt': maintenant})
                & Q(**{f('periode_fin') + '__gte': seuil_grace})
            )
        )

    if code == AbonnementOrganisation.STATUT_EXPIRE:
        return (
            Q(**{f('statut'): AbonnementOrganisation.STATUT_EXPIRE})
            | (
                Q(**{f('statut'): AbonnementOrganisation.STATUT_ESSAI})
                & Q(**{f('essai_fin') + '__isnull': False})
                & Q(**{f('essai_fin') + '__lt': maintenant})
            )
            | (
                Q(**{f('statut'): AbonnementOrganisation.STATUT_EN_RETARD})
                & (
                    Q(**{f('periode_fin') + '__isnull': True})
                    | Q(**{f('periode_fin') + '__lt': seuil_grace})
                )
            )
            | (
                Q(**{f('statut'): AbonnementOrganisation.STATUT_ACTIF})
                & Q(**{f('periode_fin') + '__isnull': False})
                & Q(**{f('periode_fin') + '__lt': seuil_grace})
            )
        )

    # Statut inconnu : pas de match
    return Q(pk__in=[])


def _echeance_info(abo, maintenant=None):
    """Date d’échéance + libellé, basés sur le statut effectif (sans mutation)."""
    maintenant = maintenant or timezone.now()
    if not abo:
        return None, '—'
    statut = _statut_effectif(abo, maintenant)
    grace = None
    if statut == AbonnementOrganisation.STATUT_ESSAI:
        dt = abo.essai_fin
    elif statut == AbonnementOrganisation.STATUT_ACTIF:
        dt = abo.periode_fin
    elif statut == AbonnementOrganisation.STATUT_EN_RETARD:
        # Fin d’accès = fin de grâce (plus lisible que periode déjà passée)
        dt = abo.fin_grace_retard() or abo.periode_fin
        grace = True
    elif statut in (
        AbonnementOrganisation.STATUT_EXPIRE,
        AbonnementOrganisation.STATUT_ANNULE,
    ):
        if abo.periode_fin and abo.essai_fin:
            dt = max(abo.periode_fin, abo.essai_fin)
        else:
            dt = abo.periode_fin or abo.essai_fin
    else:
        # prelaunch
        dt = abo.essai_fin or abo.periode_fin

    if not dt:
        return None, '—'

    if statut in (
        AbonnementOrganisation.STATUT_EXPIRE,
        AbonnementOrganisation.STATUT_ANNULE,
    ):
        if dt >= maintenant:
            return dt, 'accès coupé'
        return dt, 'dépassée'

    jours = _jours_relatifs(dt, maintenant)
    label = _libelle_jours_relatifs(jours)
    if statut == AbonnementOrganisation.STATUT_EN_RETARD and grace:
        if jours is not None and jours >= 0:
            label = f'grâce · {label}'
        else:
            label = 'dépassée'
    return dt, label


def _affichage_periodes_abo(abo, maintenant=None):
    """
    Quelle plage afficher selon le statut effectif — une seule source de vérité.
    Évite essai NOV + payant JUIL en même temps.
    """
    if not abo:
        return False, False
    s = _statut_effectif(abo, maintenant)
    if s in (
        AbonnementOrganisation.STATUT_ESSAI,
        AbonnementOrganisation.STATUT_PRELAUNCH,
    ):
        return bool(abo.essai_debut or abo.essai_fin), False
    if s in (
        AbonnementOrganisation.STATUT_ACTIF,
        AbonnementOrganisation.STATUT_EN_RETARD,
    ):
        return False, bool(abo.periode_debut or abo.periode_fin)
    if s in (
        AbonnementOrganisation.STATUT_EXPIRE,
        AbonnementOrganisation.STATUT_ANNULE,
    ):
        if abo.periode_fin or abo.periode_debut:
            return False, True
        if abo.essai_fin or abo.essai_debut:
            return True, False
        return False, False
    return False, False


def _membre_pour_ligne(user, filtre_statut='', filtre_plan='', maintenant=None):
    """
    Membership à afficher : celui qui matche statut *effectif* / plan filtrés,
    sinon le principal, sinon le premier.
    """
    membres = list(user.membres_organisations.all())
    if not membres:
        return None
    maintenant = maintenant or timezone.now()

    def _match(m):
        org = m.organisation
        try:
            abo = org.abonnement
        except AbonnementOrganisation.DoesNotExist:
            abo = None
        if filtre_statut:
            if not abo or _statut_effectif(abo, maintenant) != filtre_statut:
                return False
        if filtre_plan:
            if not abo or not abo.plan_id or abo.plan.code != filtre_plan:
                return False
        return True

    if filtre_statut or filtre_plan:
        for m in membres:
            if _match(m):
                return m

    principal = _membre_principal(user)
    if principal:
        return principal
    return membres[0]


def _derniers_paiements_par_org(org_ids):
    if not org_ids:
        return {}
    resultat = {}
    qs = (
        PaiementAbonnement.objects
        .filter(
            organisation_id__in=org_ids,
            statut=PaiementAbonnement.STATUT_REUSSI,
        )
        .order_by('organisation_id', '-paye_le', '-cree_le')
    )
    for paiement in qs:
        if paiement.organisation_id not in resultat:
            resultat[paiement.organisation_id] = paiement
    return resultat


def _lignes_utilisateurs(
    users,
    emails_verifies,
    maintenant,
    user_ids_verifies=None,
    periode_debut=None,
    periode_fin=None,
    filtre_statut='',
    filtre_plan='',
):
    users = list(users)
    org_ids = []
    membres_by_user = {}
    for user in users:
        membre = _membre_pour_ligne(
            user, filtre_statut, filtre_plan, maintenant=maintenant,
        )
        membres_by_user[user.pk] = membre
        if membre and membre.organisation_id:
            org_ids.append(membre.organisation_id)
    derniers_paiements = _derniers_paiements_par_org(org_ids)

    period_on = periode_debut is not None and periode_fin is not None
    lignes = []
    emails_verifies_l = {e.lower() for e in emails_verifies}
    user_ids_verifies = user_ids_verifies or set()
    for user in users:
        membre = membres_by_user.get(user.pk)
        org = membre.organisation if membre else None
        abo = _abonnement_of(org)
        email = (user.email or user.username or '').strip()
        nom = user.get_full_name().strip()
        role = membre.get_role_display_label() if membre else ''
        statut_aff = _statut_effectif(abo, maintenant) if abo else ''
        echeance_dt, echeance_label = _echeance_info(abo, maintenant)
        paiement = derniers_paiements.get(org.pk) if org else None
        if paiement:
            date_p = paiement.paye_le or paiement.cree_le
            paiement_montant = _format_montant(paiement.montant, paiement.devise)
            paiement_date = _date_relative(date_p, maintenant)
        else:
            date_p = None
            paiement_montant = ''
            paiement_date = ''
        email_verifie = (
            user.pk in user_ids_verifies
            or email.lower() in emails_verifies_l
        )
        inscrit_periode = False
        connecte_periode = False
        if period_on:
            dj = user.date_joined
            ll = user.last_login
            inscrit_periode = bool(
                dj and periode_debut <= dj <= periode_fin
            )
            connecte_periode = bool(
                ll and periode_debut <= ll <= periode_fin
            )
        detail_url = reverse('backoffice_user_detail', args=[user.pk])
        if org:
            detail_url = f'{detail_url}?org={org.pk}'
        lignes.append({
            'email': email,
            'nom': nom,
            'role': role,
            'actif': user.is_active and (membre.actif if membre else True),
            'email_verifie': email_verifie,
            'organisation': org,
            'telephone': (org.telephone if org else '') or '',
            'plan': abo.plan.nom if abo and abo.plan_id else '—',
            'statut': statut_aff,
            'statut_label': _statut_label(statut_aff) if statut_aff else 'Sans abo',
            'statut_db': abo.statut if abo else '',
            'echeance': echeance_dt,
            'echeance_label': echeance_label,
            'paiement_montant': paiement_montant,
            'paiement_date': paiement_date,
            'login_relatif': _date_relative(user.last_login, maintenant),
            'inscrit_relatif': _date_relative(user.date_joined, maintenant),
            'last_login': user.last_login,
            'date_joined': user.date_joined,
            'paiement_dt': date_p if paiement else None,
            'inscrit_sur_periode': inscrit_periode,
            'connecte_sur_periode': connecte_periode,
            'user_id': user.pk,
            'detail_url': detail_url,
            'org_detail_url': (
                reverse('backoffice_org_detail', args=[org.pk]) if org else ''
            ),
            'admin_user_url': f'/admin/auth/user/{user.pk}/change/',
            'admin_org_url': (
                f'/admin/comptes/organisation/{org.pk}/change/' if org else ''
            ),
            'admin_abo_url': (
                f'/admin/comptes/abonnementorganisation/{abo.pk}/change/' if abo else ''
            ),
        })
    return lignes


def _query_params(
    q,
    filtre_statut,
    filtre_plan,
    non_verifies,
    page=None,
    vue_complete=False,
    periode=None,
    du='',
    au='',
    user_scope='',
    pay_q='',
    pay_statut='',
    pay_page=None,
    pay_all=False,
    charge_q='',
    charge_categorie='',
    charge_page=None,
    charge_all=False,
):
    params = {}
    if q:
        params['q'] = q
    if filtre_statut:
        params['statut'] = filtre_statut
    if filtre_plan:
        params['plan'] = filtre_plan
    if non_verifies:
        params['non_verifies'] = '1'
    if user_scope in ('inscrits', 'connectes', 'sessions'):
        params['user_scope'] = user_scope
    if pay_q:
        params['pay_q'] = pay_q
    if pay_statut:
        params['pay_statut'] = pay_statut
    # Dates calendaires en priorité, sinon preset (sauf défaut all)
    if du or au:
        if du:
            params['du'] = du
        if au:
            params['au'] = au
    elif periode and periode not in ('', 'all'):
        # inclut les presets, y compris un custom recalculé en clé preset
        params['periode'] = periode
    if vue_complete:
        params['all'] = '1'
    elif page and int(page) > 1:
        params['page'] = page
    if pay_all:
        params['pay_all'] = '1'
    elif pay_page and int(pay_page) > 1:
        params['pay_page'] = pay_page
    if charge_q:
        params['charge_q'] = charge_q
    if charge_categorie:
        params['charge_categorie'] = charge_categorie
    if charge_all:
        params['charge_all'] = '1'
    elif charge_page and int(charge_page) > 1:
        params['charge_page'] = charge_page
    return params


def _filtres_depuis_request(request):
    q = (request.GET.get('q') or '').strip()
    filtre_statut = (request.GET.get('statut') or '').strip()
    filtre_plan = (request.GET.get('plan') or '').strip()
    non_verifies = request.GET.get('non_verifies') in ('1', 'true', 'on')
    scope_raw = (request.GET.get('user_scope') or '').strip().lower()
    user_scope = scope_raw if scope_raw in ('inscrits', 'connectes', 'sessions', 'all', '') else ''
    if user_scope == 'all':
        user_scope = ''
    pay_q = (request.GET.get('pay_q') or '').strip()
    pay_statut = (request.GET.get('pay_statut') or '').strip()
    pay_codes = {c for c, _ in PaiementAbonnement.STATUT_CHOICES}
    if pay_statut not in pay_codes:
        pay_statut = ''
    charge_q = (request.GET.get('charge_q') or '').strip()
    charge_categorie = (request.GET.get('charge_categorie') or '').strip()
    charge_codes = {c for c, _ in ChargePlateforme.CATEGORIE_CHOICES}
    if charge_categorie not in charge_codes:
        charge_categorie = ''
    return (
        q, filtre_statut, filtre_plan, non_verifies, user_scope,
        pay_q, pay_statut, charge_q, charge_categorie,
    )


def _q_date_paiement_periode(debut, fin):
    """Date effective d’un paiement (paye_le sinon cree_le) dans la période."""
    return (
        Q(paye_le__gte=debut, paye_le__lte=fin)
        | Q(paye_le__isnull=True, cree_le__gte=debut, cree_le__lte=fin)
    )


def _queryset_paiements_filtres(pay_q='', pay_statut='', debut=None, fin=None):
    """Liste des paiements d’abonnement, filtrable période / statut / recherche."""
    qs = PaiementAbonnement.objects.select_related(
        'organisation',
        'abonnement',
        'abonnement__plan',
    ).order_by('-cree_le', '-id')
    if debut is not None and fin is not None:
        qs = qs.filter(_q_date_paiement_periode(debut, fin))
    if pay_statut:
        qs = qs.filter(statut=pay_statut)
    if pay_q:
        qs = qs.filter(
            Q(organisation__nom__icontains=pay_q)
            | Q(organisation__slug__icontains=pay_q)
            | Q(organisation__email__icontains=pay_q)
            | Q(reference_externe__icontains=pay_q)
            | Q(methode__icontains=pay_q)
        )
    return qs


def _stats_paiements(qs):
    """Totaux pour l’intro de l’onglet Paiements (même base que le tableau filtré)."""
    agg = qs.aggregate(
        n=Count('id'),
        total_reussi=Sum(
            'montant',
            filter=Q(statut=PaiementAbonnement.STATUT_REUSSI),
        ),
        n_reussi=Count('id', filter=Q(statut=PaiementAbonnement.STATUT_REUSSI)),
        n_echec=Count('id', filter=Q(statut=PaiementAbonnement.STATUT_ECHEC)),
        n_attente=Count('id', filter=Q(statut=PaiementAbonnement.STATUT_EN_ATTENTE)),
        n_rembourse=Count('id', filter=Q(statut=PaiementAbonnement.STATUT_REMBOURSE)),
    )
    return {
        'n': agg.get('n') or 0,
        'total_reussi': agg.get('total_reussi'),
        'total_reussi_formate': _format_montant(agg.get('total_reussi')),
        'n_reussi': agg.get('n_reussi') or 0,
        'n_echec': agg.get('n_echec') or 0,
        'n_attente': agg.get('n_attente') or 0,
        'n_rembourse': agg.get('n_rembourse') or 0,
    }


def _lignes_paiements(paiements, maintenant=None):
    """Lignes affichables pour la table Paiements."""
    maintenant = maintenant or timezone.now()
    lignes = []
    for p in paiements:
        org = p.organisation
        abo = p.abonnement
        date_eff = p.paye_le or p.cree_le
        lignes.append({
            'id': p.pk,
            'org_nom': (org.nom or org.slug) if org else '—',
            'org_id': org.pk if org else None,
            'plan': abo.plan.nom if abo and abo.plan_id else '—',
            'montant': _format_montant(p.montant, p.devise or 'XOF'),
            'montant_raw': p.montant,
            'devise': p.devise or 'XOF',
            'statut': p.statut,
            'statut_label': p.get_statut_display(),
            'methode': p.methode or '—',
            'reference': p.reference_externe or '',
            'date': date_eff,
            'date_label': _format_dt_fr(date_eff),
            'date_relatif': _date_relative(date_eff, maintenant) if date_eff else '—',
            'periode_couverte': (
                f'{_format_dt_fr(p.periode_couverte_debut)} → {_format_dt_fr(p.periode_couverte_fin)}'
                if p.periode_couverte_debut or p.periode_couverte_fin
                else ''
            ),
            'admin_url': f'/admin/comptes/paiementabonnement/{p.pk}/change/',
            'admin_org_url': (
                f'/admin/comptes/organisation/{org.pk}/change/' if org else ''
            ),
            'detail_url': (
                reverse('backoffice_org_detail', args=[org.pk]) if org else ''
            ),
            'href_users': _users_href({'q': org.nom if org else ''}) if org else '#utilisateurs',
        })
    return lignes


def _q_users_activite_periode(debut, fin, prefix=''):
    """Utilisateurs inscrits ou connectés sur la période.
    prefix='' pour User, 'utilisateur__' pour MembreOrganisation, etc.
    """
    p = prefix or ''
    return (
        Q(**{f'{p}date_joined__gte': debut, f'{p}date_joined__lte': fin})
        | Q(**{f'{p}last_login__gte': debut, f'{p}last_login__lte': fin})
    )


def _q_users_scope_periode(debut, fin, scope=''):
    """
    Filtre période pour la liste utilisateurs.
    scope '' | 'all' → inscrits OU connectés
    'inscrits' → date_joined dans la période
    'connectes' → last_login dans la période
    """
    if scope == 'inscrits':
        return Q(date_joined__gte=debut, date_joined__lte=fin)
    if scope == 'connectes':
        return Q(last_login__gte=debut, last_login__lte=fin)
    return _q_users_activite_periode(debut, fin)


def _org_ids_actives_periode(debut, fin):
    """Orgs touchées par la période (créées, paiement, ou membre actif sur la période)."""
    ids = set(
        Organisation.objects.filter(
            cree_le__gte=debut,
            cree_le__lte=fin,
        ).values_list('pk', flat=True)
    )
    ids.update(
        MembreOrganisation.objects.filter(
            _q_users_activite_periode(debut, fin, prefix='utilisateur__')
        ).values_list('organisation_id', flat=True)
    )
    ids.update(
        PaiementAbonnement.objects.filter(
            statut=PaiementAbonnement.STATUT_REUSSI,
        ).filter(
            Q(paye_le__gte=debut, paye_le__lte=fin)
            | Q(paye_le__isnull=True, cree_le__gte=debut, cree_le__lte=fin)
        ).values_list('organisation_id', flat=True)
    )
    ids.discard(None)
    return ids


def _compter_par_statut_abo(abo_qs=None, maintenant=None):
    """
    Répartition abonnements par statut *effectif* (dates + grâce),
    pas le simple COUNT(statut) en base.
    Retourne aussi le MRR (somme des prix des plans effectivement actifs).
    """
    maintenant = maintenant or timezone.now()
    qs = abo_qs if abo_qs is not None else AbonnementOrganisation.objects.all()
    # select_related(None) : évite le conflit si l’appelant a déjà join organisation
    # + only() qui defers ce champ.
    qs = (
        qs.select_related(None)
        .select_related('plan')
        .only(
            'id',
            'statut',
            'essai_fin',
            'periode_fin',
            'plan_id',
            'plan__prix_mensuel',
        )
    )
    par_statut = {
        code: 0 for code, _label in AbonnementOrganisation.STATUT_CHOICES
    }
    mrr = Decimal('0')
    for abo in qs.iterator(chunk_size=500):
        code = _statut_effectif(abo, maintenant) or abo.statut
        if code not in par_statut:
            par_statut[code] = 0
        par_statut[code] += 1
        if (
            code == AbonnementOrganisation.STATUT_ACTIF
            and abo.plan_id
            and abo.plan.prix_mensuel
        ):
            mrr += abo.plan.prix_mensuel
    return par_statut, mrr


def _ids_users_session_en_cours():
    """IDs utilisateurs ayant une session Django non expirée (connectés maintenant)."""
    maintenant = timezone.now()
    ids = set()
    qs = Session.objects.filter(expire_date__gte=maintenant).only('session_data')
    for session in qs.iterator(chunk_size=200):
        try:
            data = session.get_decoded()
        except Exception:
            continue
        uid = data.get('_auth_user_id')
        if uid is None:
            continue
        try:
            ids.add(int(uid))
        except (TypeError, ValueError):
            continue
    return ids


def _compter_users_session_en_cours():
    """Nombre de comptes distincts avec session en cours."""
    ids = _ids_users_session_en_cours()
    if not ids:
        return 0
    return User.objects.filter(pk__in=ids, is_active=True).count()


def _paiements_ok_periode(debut=None, fin=None):
    qs = PaiementAbonnement.objects.filter(statut=PaiementAbonnement.STATUT_REUSSI)
    if debut is not None and fin is not None:
        qs = qs.filter(
            Q(paye_le__gte=debut, paye_le__lte=fin)
            | Q(paye_le__isnull=True, cree_le__gte=debut, cree_le__lte=fin)
        )
    return qs


def _queryset_utilisateurs_filtres(
    q,
    filtre_statut,
    filtre_plan,
    non_verifies,
    periode_debut=None,
    periode_fin=None,
    user_scope='',
    maintenant=None,
):
    membres_prefetch = Prefetch(
        'membres_organisations',
        queryset=MembreOrganisation.objects.select_related(
            'organisation',
            'organisation__abonnement',
            'organisation__abonnement__plan',
        ),
    )
    periode_on = periode_debut is not None and periode_fin is not None
    sessions_on = user_scope == 'sessions'
    if periode_on or sessions_on:
        # Activité récente en premier
        order = ('-last_login', '-date_joined', '-id')
    else:
        order = ('-date_joined', '-id')

    users_list = (
        User.objects
        .select_related('profil')
        .prefetch_related(membres_prefetch)
        .order_by(*order)
    )

    # En ligne maintenant (sessions Django) — indépendant de la période calendaire
    if sessions_on:
        session_ids = _ids_users_session_en_cours()
        users_list = users_list.filter(pk__in=session_ids or [-1])
    elif periode_on:
        users_list = users_list.filter(
            _q_users_scope_periode(periode_debut, periode_fin, user_scope)
        )

    if q:
        users_list = users_list.filter(
            Q(email__icontains=q)
            | Q(username__icontains=q)
            | Q(first_name__icontains=q)
            | Q(last_name__icontains=q)
            | Q(membres_organisations__organisation__nom__icontains=q)
            | Q(membres_organisations__organisation__slug__icontains=q)
        ).distinct()

    if filtre_statut:
        # Statut effectif (dates + grâce) — aligné KPI / pastilles
        q_eff = _q_statut_abonnement_effectif(
            filtre_statut,
            maintenant=maintenant or timezone.now(),
            prefix='membres_organisations__organisation__abonnement__',
        )
        users_list = users_list.filter(q_eff).distinct()

    if filtre_plan:
        users_list = users_list.filter(
            membres_organisations__organisation__abonnement__plan__code=filtre_plan,
        ).distinct()

    emails_verifies = _emails_verifies()
    user_ids_verifies = _ids_users_email_verifie()
    if non_verifies:
        users_list = users_list.exclude(pk__in=user_ids_verifies).annotate(
            email_l=Lower('email'),
        ).exclude(email_l__in=emails_verifies)

    return users_list, emails_verifies, user_ids_verifies


def _stats_liste_utilisateurs(users_qs, periode_debut=None, periode_fin=None):
    """Compteurs pour l’intro onglet Utilisateurs (même base que le tableau)."""
    total = users_qs.count()
    n_actifs = users_qs.filter(is_active=True).count()
    period_on = periode_debut is not None and periode_fin is not None
    if period_on:
        n_inscrits = users_qs.filter(
            date_joined__gte=periode_debut,
            date_joined__lte=periode_fin,
        ).count()
        n_connectes = users_qs.filter(
            last_login__gte=periode_debut,
            last_login__lte=periode_fin,
        ).count()
    else:
        n_inscrits = total
        n_connectes = users_qs.exclude(last_login__isnull=True).count()
    return {
        'total': total,
        'actifs': n_actifs,
        'inscriptions': n_inscrits,
        'connexions': n_connectes,
        'period_on': period_on,
    }


def _dt_excel(dt):
    if not dt:
        return ''
    local = timezone.localtime(dt) if timezone.is_aware(dt) else dt
    return local.strftime('%d/%m/%Y %H:%M')


def _date_excel(dt):
    if not dt:
        return ''
    local = timezone.localtime(dt) if timezone.is_aware(dt) else dt
    return local.strftime('%d/%m/%Y')


def _sondages_backoffice(page_number=1, per_page=2):
    from finances.models import Sondage, SondageOption

    option_qs = SondageOption.objects.annotate(
        nb_reponses=Count('reponses')
    ).order_by('ordre', 'pk')
    polls_qs = (
        Sondage.objects.select_related('cree_par')
        .prefetch_related(Prefetch('options', queryset=option_qs))
        .annotate(nb_reponses=Count('reponses', distinct=True))
        .order_by('-cree_le', '-pk')
    )
    paginator = Paginator(polls_qs, per_page)
    page_obj = paginator.get_page(page_number)
    current = page_obj.number
    total_pages = paginator.num_pages
    page_numbers = list(range(max(1, current - 2), min(total_pages, current + 2) + 1))

    rows = []
    for poll in page_obj.object_list:
        total = int(poll.nb_reponses or 0)
        options = []
        for option in poll.options.all():
            count = int(option.nb_reponses or 0)
            options.append({
                'texte': option.texte,
                'nb_reponses': count,
                'pourcentage': round((count * 100 / total), 1) if total else 0,
            })
        rows.append({
            'id': poll.pk,
            'question': poll.question,
            'actif': poll.actif,
            'cree_le': poll.cree_le,
            'nb_reponses': total,
            'options': options,
        })
    return {
        'rows': rows,
        'page_obj': page_obj,
        'page_numbers': page_numbers,
        'total': paginator.count,
    }


@backoffice_required
def backoffice_dashboard(request):
    maintenant = timezone.now()
    partial = _partial_kind(request)
    light_partial = partial in ('users', 'payments', 'finances', 'polls')
    periode = _periode_depuis_request(request)
    periode_key = periode['key']
    periode_debut = periode['debut']
    periode_fin = periode['fin']
    periode_label = periode['label']
    filtre_actif = periode['filtre_actif']
    charts_debut = periode.get('charts_debut', periode_debut)
    charts_fin = periode.get('charts_fin', periode_fin)
    if charts_debut is None and charts_fin is None:
        charts_debut = _debut_historique_plateforme(timezone.localtime(maintenant))
        charts_fin = maintenant
    du_val = periode.get('du') or ''
    au_val = periode.get('au') or ''

    q, filtre_statut, filtre_plan, non_verifies, user_scope, pay_q, pay_statut, charge_q, charge_categorie = _filtres_depuis_request(request)

    # ── STOCK plateforme (indépendant du filtre temporel) ──────────────────
    # État courant abonnements / MRR / expirés : ce n’est pas un flux de période.
    abo_qs = AbonnementOrganisation.objects.select_related('organisation', 'plan')
    par_statut, mrr = _compter_par_statut_abo(abo_qs, maintenant=maintenant)
    acces_ouverts = (
        par_statut[AbonnementOrganisation.STATUT_PRELAUNCH]
        + par_statut[AbonnementOrganisation.STATUT_ESSAI]
        + par_statut[AbonnementOrganisation.STATUT_ACTIF]
        + par_statut[AbonnementOrganisation.STATUT_EN_RETARD]
    )
    mrr_formate = _format_montant(mrr)

    users_all = User.objects.all()
    nb_comptes = users_all.count()
    nb_comptes_actifs = users_all.filter(is_active=True).count()
    nb_orgs = Organisation.objects.count()
    sans_org = users_all.filter(membres_organisations__isnull=True).count()
    connectes_maintenant = _compter_users_session_en_cours()

    # Cumul historique des paiements réussis (toujours global)
    paiements_cumul = _paiements_ok_periode().aggregate(
        total=Sum('montant'), n=Count('id'),
    )
    total_recu_cumul = paiements_cumul.get('total')
    n_paiements_cumul = paiements_cumul.get('n') or 0

    # ── FLUX sur la période (ou = tout si filtre inactif) ───────────────────
    if filtre_actif and periode_debut and periode_fin:
        debut, fin = periode_debut, periode_fin
        inscriptions = users_all.filter(
            date_joined__gte=debut, date_joined__lte=fin,
        ).count()
        nouvelles_orgs = Organisation.objects.filter(
            cree_le__gte=debut, cree_le__lte=fin,
        ).count()
        paiements_flux = _paiements_ok_periode(debut, fin).aggregate(
            total=Sum('montant'), n=Count('id'),
        )
        hint_flux = periode_label
    else:
        inscriptions = nb_comptes
        nouvelles_orgs = nb_orgs
        paiements_flux = paiements_cumul
        hint_flux = 'toutes les données'

    encaisse = paiements_flux.get('total')
    n_encaisse = paiements_flux.get('n') or 0

    if filtre_actif and periode_debut and periode_fin:
        charges_flux = _charges_periode(periode_debut, periode_fin).aggregate(
            total=Sum('montant'), n=Count('id'),
        )
    else:
        charges_flux = ChargePlateforme.objects.aggregate(
            total=Sum('montant'), n=Count('id'),
        )
    total_charges_flux = charges_flux.get('total') or Decimal('0')
    n_charges_flux = charges_flux.get('n') or 0
    encaisse_val = encaisse or Decimal('0')
    resultat_net_flux = encaisse_val - total_charges_flux

    sante_financiere = {
        'encaisse_formate': _format_montant(encaisse),
        'charges_formate': _format_montant(total_charges_flux),
        'resultat_formate': _format_montant(resultat_net_flux),
        'resultat_positif': resultat_net_flux >= 0,
        'n_charges': n_charges_flux,
        'hint_flux': hint_flux,
    }

    # ── KPI Vue d'ensemble — 8 cartes stables ─────────────────────────────
    # STOCK : Comptes, Entreprises, Abonnements, Revenu prévu, Connectés, Expirés
    # FLUX  : Total reçu (cumul all-time) vs Encaissé (période)
    kpis = [
        {
            'label': 'Comptes',
            'valeur': nb_comptes,
            'hint': (
                f'{inscriptions} inscription{"s" if inscriptions != 1 else ""} · {hint_flux}'
                if filtre_actif
                else f'{nb_comptes_actifs} actif{"s" if nb_comptes_actifs != 1 else ""}'
            ),
            'variant': 'balance',
        },
        {
            'label': 'Entreprises',
            'valeur': nb_orgs,
            'hint': (
                f'{nouvelles_orgs} nouvelle{"s" if nouvelles_orgs != 1 else ""} · {hint_flux}'
                if filtre_actif
                else (
                    f'{sans_org} sans organisation'
                    if sans_org
                    else '0 sans organisation'
                )
            ),
            'variant': 'remaining' if sans_org else 'balance',
        },
        {
            'label': 'Abonnements',
            'valeur': acces_ouverts,
            'hint': (
                f'{par_statut[AbonnementOrganisation.STATUT_ESSAI]} essai · '
                f'{par_statut[AbonnementOrganisation.STATUT_ACTIF]} payants'
            ),
            'variant': 'income',
        },
        {
            'label': 'Revenu prévu',
            'valeur': mrr_formate,
            'hint': f'{par_statut[AbonnementOrganisation.STATUT_ACTIF]} abonnement(s) payant(s)',
            'variant': 'income',
        },
        {
            'label': 'Total reçu',
            'valeur': _format_montant(total_recu_cumul),
            'hint': f'{n_paiements_cumul} paiement(s) cumulés',
            'variant': 'income',
        },
        {
            'label': 'En ligne',
            'valeur': connectes_maintenant,
            'hint': 'sessions ouvertes maintenant · pas le filtre période',
            'variant': 'income' if connectes_maintenant else 'remaining',
            'href': None,  # rempli après construction des query params
        },
        {
            'label': 'Encaissé',
            'valeur': _format_montant(encaisse),
            'hint': f'{n_encaisse} paiement(s) · {hint_flux}',
            'variant': 'income',
        },
        {
            'label': 'Expirés',
            'valeur': par_statut[AbonnementOrganisation.STATUT_EXPIRE],
            'hint': f'{par_statut[AbonnementOrganisation.STATUT_ANNULE]} annulé(s)',
            'variant': (
                'expense'
                if par_statut[AbonnementOrganisation.STATUT_EXPIRE]
                else 'balance'
            ),
        },
    ]

    # ── Liste utilisateurs : période (activité) + recherche + abo ──────────
    # partial=payments|finances : on saute l’annuaire (soft-nav listes)
    skip_users = partial in ('payments', 'finances', 'polls')
    skip_payments = partial in ('users', 'finances', 'polls')
    skip_finances = partial in ('users', 'payments', 'polls')

    p_debut, p_fin, scope_effectif = _params_liste_utilisateurs(
        filtre_actif, periode_debut, periode_fin, user_scope,
    )
    badge_debut = periode_debut if filtre_actif else None
    badge_fin = periode_fin if filtre_actif else None
    page = None
    page_numbers = []
    lignes = []
    users_totaux = {
        'total': 0, 'actifs': 0, 'inscriptions': 0, 'connexions': 0,
    }
    vue_complete = request.GET.get('all') in ('1', 'true', 'on')
    nb_resultats = 0

    if not skip_users:
        users_list, emails_verifies, user_ids_verifies = _queryset_utilisateurs_filtres(
            q,
            filtre_statut,
            filtre_plan,
            non_verifies,
            periode_debut=p_debut,
            periode_fin=p_fin,
            user_scope=scope_effectif,
            maintenant=maintenant,
        )
        users_totaux = _stats_liste_utilisateurs(users_list, badge_debut, badge_fin)
        nb_resultats = users_totaux['total']

        if vue_complete:
            lignes = _lignes_utilisateurs(
                users_list,
                emails_verifies,
                maintenant,
                user_ids_verifies=user_ids_verifies,
                periode_debut=badge_debut,
                periode_fin=badge_fin,
                filtre_statut=filtre_statut,
                filtre_plan=filtre_plan,
            )
        else:
            paginator = Paginator(users_list, 5)
            page = paginator.get_page(request.GET.get('page') or 1)
            lignes = _lignes_utilisateurs(
                page.object_list,
                emails_verifies,
                maintenant,
                user_ids_verifies=user_ids_verifies,
                periode_debut=badge_debut,
                periode_fin=badge_fin,
                filtre_statut=filtre_statut,
                filtre_plan=filtre_plan,
            )
            current = page.number
            total_pages = paginator.num_pages
            window_start = max(1, current - 2)
            window_end = min(total_pages, current + 2)
            page_numbers = list(range(window_start, window_end + 1))
            nb_resultats = paginator.count

    # ── Liste paiements abonnement (période + statut + recherche) ──────────
    p_debut_pay = periode_debut if filtre_actif else None
    p_fin_pay = periode_fin if filtre_actif else None
    pay_page = None
    pay_page_numbers = []
    lignes_paiements = []
    pays_totaux = {
        'n': 0, 'n_reussi': 0, 'n_echec': 0, 'n_attente': 0, 'n_rembourse': 0,
        'total_reussi': None, 'total_reussi_formate': _format_montant(None),
    }
    pays_totaux_base = dict(pays_totaux)
    pay_vue_complete = request.GET.get('pay_all') in ('1', 'true', 'on')
    nb_paiements = 0

    if not skip_payments:
        pays_base = _queryset_paiements_filtres(
            pay_q=pay_q,
            pay_statut='',
            debut=p_debut_pay,
            fin=p_fin_pay,
        )
        pays_totaux_base = _stats_paiements(pays_base)
        pays_list = pays_base.filter(statut=pay_statut) if pay_statut else pays_base
        pays_totaux = _stats_paiements(pays_list) if pay_statut else pays_totaux_base
        nb_paiements = pays_totaux['n']
        if pay_vue_complete:
            lignes_paiements = _lignes_paiements(pays_list, maintenant)
        else:
            pay_paginator = Paginator(pays_list, 10)
            pay_page = pay_paginator.get_page(request.GET.get('pay_page') or 1)
            lignes_paiements = _lignes_paiements(pay_page.object_list, maintenant)
            pcur = pay_page.number
            ptotal = pay_paginator.num_pages
            pay_page_numbers = list(range(max(1, pcur - 2), min(ptotal, pcur + 2) + 1))
            nb_paiements = pay_paginator.count

    # ── Liste charges plateforme (période + catégorie + recherche) ─────────
    p_debut_ch = periode_debut if filtre_actif else None
    p_fin_ch = periode_fin if filtre_actif else None
    charge_page = None
    charge_page_numbers = []
    lignes_charges = []
    charges_totaux = {'n': 0, 'total': Decimal('0'), 'total_formate': _format_montant(None)}
    charges_totaux_base = dict(charges_totaux)
    charge_vue_complete = request.GET.get('charge_all') in ('1', 'true', 'on')
    nb_charges = 0
    charges_base = ChargePlateforme.objects.none()

    if not skip_finances:
        charges_base = _queryset_charges_filtres(
            charge_q=charge_q,
            charge_categorie='',
            debut=p_debut_ch,
            fin=p_fin_ch,
        )
        charges_totaux_base = _stats_charges(charges_base)
        charges_list = (
            charges_base.filter(categorie=charge_categorie)
            if charge_categorie
            else charges_base
        )
        charges_totaux = (
            _stats_charges(charges_list) if charge_categorie else charges_totaux_base
        )
        nb_charges = charges_totaux['n']
        if charge_vue_complete:
            lignes_charges = _lignes_charges(charges_list)
        else:
            charge_paginator = Paginator(charges_list, 10)
            charge_page = charge_paginator.get_page(request.GET.get('charge_page') or 1)
            lignes_charges = _lignes_charges(charge_page.object_list)
            ccur = charge_page.number
            ctotal = charge_paginator.num_pages
            charge_page_numbers = list(range(max(1, ccur - 2), min(ctotal, ccur + 2) + 1))
            nb_charges = charge_paginator.count

    q_kwargs = dict(
        q=q,
        filtre_statut=filtre_statut,
        filtre_plan=filtre_plan,
        non_verifies=non_verifies,
        periode=periode_key,
        du=du_val,
        au=au_val,
        user_scope=scope_effectif,
        pay_q=pay_q,
        pay_statut=pay_statut,
        charge_q=charge_q,
        charge_categorie=charge_categorie,
    )
    query_suffix = urlencode(
        _query_params(
            **q_kwargs, vue_complete=vue_complete,
            pay_all=pay_vue_complete, charge_all=charge_vue_complete,
        )
    )
    query_suffix_pages = urlencode(
        _query_params(**q_kwargs, vue_complete=False, pay_all=False, charge_all=False)
    )
    query_suffix_sans_statut = urlencode(
        _query_params(
            q=q,
            filtre_statut='',
            filtre_plan=filtre_plan,
            non_verifies=non_verifies,
            vue_complete=False,
            periode=periode_key,
            du=du_val,
            au=au_val,
            user_scope=scope_effectif,
            pay_q=pay_q,
            pay_statut=pay_statut,
        )
    )
    # Reset des filtres liste (q / statut / plan / non_vérifiés / scope) en gardant la période
    query_suffix_periode_seule = urlencode(
        _query_params(
            q='',
            filtre_statut='',
            filtre_plan='',
            non_verifies=False,
            vue_complete=False,
            periode=periode_key,
            du=du_val,
            au=au_val,
            user_scope='',
            pay_q=pay_q,
            pay_statut=pay_statut,
        )
    )
    # Liens onglet paiements (pagination / reset filtres pay / chips statut)
    query_suffix_pay_pages = urlencode(
        _query_params(**q_kwargs, vue_complete=vue_complete, pay_all=False)
    )
    query_suffix_pay_clear = urlencode(
        _query_params(
            q=q,
            filtre_statut=filtre_statut,
            filtre_plan=filtre_plan,
            non_verifies=non_verifies,
            vue_complete=False,
            periode=periode_key,
            du=du_val,
            au=au_val,
            user_scope=scope_effectif,
            pay_q='',
            pay_statut='',
            pay_all=False,
        )
    )
    query_suffix_charge_pages = urlencode(
        _query_params(**q_kwargs, vue_complete=vue_complete, pay_all=pay_vue_complete, charge_all=False)
    )
    query_suffix_charge_clear = urlencode(
        _query_params(
            q=q,
            filtre_statut=filtre_statut,
            filtre_plan=filtre_plan,
            non_verifies=non_verifies,
            vue_complete=False,
            periode=periode_key,
            du=du_val,
            au=au_val,
            user_scope=scope_effectif,
            pay_q=pay_q,
            pay_statut=pay_statut,
            charge_q='',
            charge_categorie='',
            charge_all=False,
        )
    )
    query_suffix_sondage_pages = urlencode(
        _query_params(
            **q_kwargs,
            vue_complete=vue_complete,
            pay_all=pay_vue_complete,
            charge_all=charge_vue_complete,
        )
    )
    if light_partial and partial != 'polls':
        sondages_data = {
            'rows': [],
            'page_obj': None,
            'page_numbers': [],
            'total': 0,
        }
    else:
        sondages_data = _sondages_backoffice(
            page_number=request.GET.get('sondage_page') or 1,
        )

    def _href_pay_statut(code=''):
        qs = urlencode(
            _query_params(
                q=q,
                filtre_statut=filtre_statut,
                filtre_plan=filtre_plan,
                non_verifies=non_verifies,
                vue_complete=False,
                periode=periode_key,
                du=du_val,
                au=au_val,
                user_scope=scope_effectif,
                pay_q=pay_q,
                pay_statut=code,
                pay_all=False,
            )
        )
        return f'?{qs}#paiements' if qs else '#paiements'

    pay_statut_chips = [
        {
            'code': '',
            'label': 'Tous',
            'n': pays_totaux_base['n'],
            'href': _href_pay_statut(''),
            'active': not pay_statut,
        },
        {
            'code': PaiementAbonnement.STATUT_REUSSI,
            'label': 'Réussis',
            'n': pays_totaux_base['n_reussi'],
            'href': _href_pay_statut(PaiementAbonnement.STATUT_REUSSI),
            'active': pay_statut == PaiementAbonnement.STATUT_REUSSI,
        },
        {
            'code': PaiementAbonnement.STATUT_ECHEC,
            'label': 'Échecs',
            'n': pays_totaux_base['n_echec'],
            'href': _href_pay_statut(PaiementAbonnement.STATUT_ECHEC),
            'active': pay_statut == PaiementAbonnement.STATUT_ECHEC,
        },
        {
            'code': PaiementAbonnement.STATUT_EN_ATTENTE,
            'label': 'En attente',
            'n': pays_totaux_base['n_attente'],
            'href': _href_pay_statut(PaiementAbonnement.STATUT_EN_ATTENTE),
            'active': pay_statut == PaiementAbonnement.STATUT_EN_ATTENTE,
        },
        {
            'code': PaiementAbonnement.STATUT_REMBOURSE,
            'label': 'Remboursés',
            'n': pays_totaux_base['n_rembourse'],
            'href': _href_pay_statut(PaiementAbonnement.STATUT_REMBOURSE),
            'active': pay_statut == PaiementAbonnement.STATUT_REMBOURSE,
        },
    ]

    def _href_charge_categorie(code=''):
        qs = urlencode(
            _query_params(
                q=q,
                filtre_statut=filtre_statut,
                filtre_plan=filtre_plan,
                non_verifies=non_verifies,
                vue_complete=False,
                periode=periode_key,
                du=du_val,
                au=au_val,
                user_scope=scope_effectif,
                pay_q=pay_q,
                pay_statut=pay_statut,
                charge_q=charge_q,
                charge_categorie=code,
                charge_all=False,
            )
        )
        return f'?{qs}#finances' if qs else '#finances'

    cat_counts = {}
    if not skip_finances:
        for row in charges_base.values('categorie').annotate(n=Count('id')):
            cat_counts[row['categorie']] = row['n']

    charge_categorie_chips = [
        {
            'code': '',
            'label': 'Toutes',
            'n': charges_totaux_base['n'],
            'href': _href_charge_categorie(''),
            'active': not charge_categorie,
        },
    ]
    for code, label in ChargePlateforme.CATEGORIE_CHOICES:
        charge_categorie_chips.append({
            'code': code,
            'label': label,
            'n': cat_counts.get(code, 0),
            'href': _href_charge_categorie(code),
            'active': charge_categorie == code,
        })

    sessions_q = urlencode(
        _query_params(
            q=q,
            filtre_statut=filtre_statut,
            filtre_plan=filtre_plan,
            non_verifies=non_verifies,
            vue_complete=False,
            periode=periode_key,
            du=du_val,
            au=au_val,
            user_scope='sessions',
            pay_q=pay_q,
            pay_statut=pay_statut,
        )
    )
    for kpi in kpis:
        if kpi.get('label') == 'En ligne':
            kpi['href'] = (
                f'?{sessions_q}#utilisateurs' if sessions_q else '?user_scope=sessions#utilisateurs'
            )
            break

    total_abo = sum(par_statut.values()) or 1
    # Partials listes : éviter graphiques / alertes (coût serveur)
    if light_partial:
        charts_data = {
            'labels': [],
            'revenus_mois': [],
            'connectes_mois': [],
            'inscrits_cumul': [],
            'unit_label': 'mois',
            'granularite': 'month',
            'totaux': {},
        }
        geo_data = {'pays': {'labels': [], 'values': []}, 'villes': {'labels': [], 'values': []}}
        alertes = {
            'total': 0,
            'horizon_jours': 7,
            'echecs_jours': 7,
            'expire_bientot': {'n': 0, 'items': []},
            'en_retard': {'n': 0, 'items': [], 'href_all': ''},
            'echecs': {'n': 0, 'items': [], 'href_all': ''},
        }
    else:
        charts_data = _graphiques_plateforme(debut=charts_debut, fin=charts_fin)
        geo_data = _repartition_geo_utilisateurs(
            debut=periode_debut if filtre_actif else None,
            fin=periode_fin if filtre_actif else None,
        )
        alertes = _alertes_abonnements(maintenant=maintenant)
        # Lien alertes échecs → onglet paiements filtrés
        href_echecs_pay = _href_pay_statut(PaiementAbonnement.STATUT_ECHEC)
        alertes['echecs']['href_all'] = href_echecs_pay

    context = {
        'kpis': kpis,
        'alertes': alertes,
        'lignes': lignes,
        'page_obj': page,
        'page_numbers': page_numbers,
        'vue_complete': vue_complete,
        'q': q,
        'filtre_statut': filtre_statut,
        'filtre_plan': filtre_plan,
        'non_verifies': non_verifies,
        'periode': periode_key,
        'periode_label': periode_label,
        'periode_mode': periode.get('mode') or 'range',
        'periode_du': du_val,
        'periode_au': au_val,
        'periode_filtre_actif': filtre_actif,
        'periodes_presets': PERIODE_PRESETS,
        'query_suffix': query_suffix,
        'query_suffix_pages': query_suffix_pages,
        'query_suffix_sans_statut': query_suffix_sans_statut,
        'query_suffix_periode_seule': query_suffix_periode_seule,
        'query_suffix_pay_pages': query_suffix_pay_pages,
        'query_suffix_pay_clear': query_suffix_pay_clear,
        'user_scope': scope_effectif,
        'users_totaux': users_totaux,
        'pay_q': pay_q,
        'pay_statut': pay_statut,
        'lignes_paiements': lignes_paiements,
        'pay_page_obj': pay_page,
        'pay_page_numbers': pay_page_numbers,
        'pay_vue_complete': pay_vue_complete,
        'nb_paiements': nb_paiements,
        'pays_totaux': pays_totaux,
        'pays_totaux_base': pays_totaux_base,
        'pay_statut_chips': pay_statut_chips,
        'pay_statuts_choices': PaiementAbonnement.STATUT_CHOICES,
        'sante_financiere': sante_financiere,
        'charge_q': charge_q,
        'charge_categorie': charge_categorie,
        'lignes_charges': lignes_charges,
        'charge_page_obj': charge_page,
        'charge_page_numbers': charge_page_numbers,
        'charge_vue_complete': charge_vue_complete,
        'nb_charges': nb_charges,
        'charges_totaux': charges_totaux,
        'charges_totaux_base': charges_totaux_base,
        'charge_categorie_chips': charge_categorie_chips,
        'charge_categories_choices': ChargePlateforme.CATEGORIE_CHOICES,
        'query_suffix_charge_pages': query_suffix_charge_pages,
        'query_suffix_charge_clear': query_suffix_charge_clear,
        'query_suffix_charge_export': urlencode(
            _query_params(
                **q_kwargs,
                vue_complete=charge_vue_complete,
                pay_all=pay_vue_complete,
                charge_all=charge_vue_complete,
            )
        ),
        'charge_date_defaut': timezone.localdate().isoformat(),
        'statuts_choices': AbonnementOrganisation.STATUT_CHOICES,
        'plans_choices': PlanAbonnement.objects.order_by('ordre', 'code'),
        'par_statut': [
            {
                'code': code,
                'label': label,
                'aide': STATUT_ABO_AIDE.get(code, ''),
                'n': par_statut[code],
                'pct': round(100 * par_statut[code] / total_abo) if total_abo else 0,
            }
            for code, label in AbonnementOrganisation.STATUT_CHOICES
        ],
        'mrr_formate': mrr_formate,
        'nb_resultats': nb_resultats,
        'nb_comptes': nb_comptes,
        'connectes_maintenant': connectes_maintenant,
        'maintenant': maintenant,
        'charts_unit_label': charts_data.get('unit_label') or 'mois',
        'charts_granularite': charts_data.get('granularite') or 'month',
        'stats_totaux': charts_data.get('totaux') or {},
        'geo_utilisateurs': geo_data,
        'acces_backoffice': _liste_acces_backoffice(request.user),
        'lancement_bo': {
            'nb_prelaunch': par_statut[AbonnementOrganisation.STATUT_PRELAUNCH],
            'disponible': par_statut[AbonnementOrganisation.STATUT_PRELAUNCH] > 0,
            'deja_lance': AbonnementOrganisation.objects.filter(
                lancement_applique_le__isnull=False,
            ).exists(),
            'date_effective': AbonnementOrganisation.date_lancement_effective(),
            'jours_essai': AbonnementOrganisation.duree_essai().days,
            'phrase_confirm': LANCEMENT_CONFIRM_PHRASE,
        },
        'notif_broadcast': {
            'nb_destinataires': (
                MembreOrganisation.objects.filter(
                    actif=True,
                    utilisateur__is_active=True,
                )
                .values('utilisateur_id')
                .distinct()
                .count()
            ),
        },
        'sondages': sondages_data['rows'],
        'sondage_page_obj': sondages_data['page_obj'],
        'sondage_page_numbers': sondages_data['page_numbers'],
        'nb_sondages': sondages_data['total'],
        'query_suffix_sondage_pages': query_suffix_sondage_pages,
        'prolongation_bo': {
            'nb_abonnements': AbonnementOrganisation.objects.count(),
            'jours_defaut': 30,
            'phrase_confirm': PROLONGER_TOUS_CONFIRM_PHRASE,
        },
        'charts_json': json.dumps(
            {
                **charts_data,
                'geo': geo_data,
                'statuts': {
                    'labels': [label for _code, label in AbonnementOrganisation.STATUT_CHOICES],
                    'values': [par_statut[code] for code, _label in AbonnementOrganisation.STATUT_CHOICES],
                    'codes': [code for code, _label in AbonnementOrganisation.STATUT_CHOICES],
                },
                'periode': periode_key,
                'periode_label': periode_label,
                'filtre_actif': filtre_actif,
            },
            ensure_ascii=False,
        ),
    }

    # ── Partials / soft-nav ────────────────────────────────────────────────
    if partial == 'users':
        html = render_to_string(
            'backoffice/partials/panel_utilisateurs.html', context, request=request
        )
        return HttpResponse(html)

    if partial == 'payments':
        html = render_to_string(
            'backoffice/partials/panel_paiements.html', context, request=request
        )
        return HttpResponse(html)

    if partial == 'finances':
        html = render_to_string(
            'backoffice/partials/panel_finances.html', context, request=request
        )
        return HttpResponse(html)

    if partial == 'polls':
        html = render_to_string(
            'backoffice/partials/panel_sondages.html', context, request=request
        )
        return HttpResponse(html)

    if partial == 'refresh':
        payload = {
            'ok': True,
            'periode_label': periode_label,
            'periode': periode_key,
            'periode_mode': periode.get('mode') or 'range',
            'periode_du': du_val,
            'periode_au': au_val,
            'html': {
                'panel-vue': render_to_string(
                    'backoffice/partials/panel_vue.html', context, request=request
                ),
                'panel-stats': render_to_string(
                    'backoffice/partials/panel_stats.html', context, request=request
                ),
                'panel-utilisateurs': render_to_string(
                    'backoffice/partials/panel_utilisateurs.html', context, request=request
                ),
                'panel-paiements': render_to_string(
                    'backoffice/partials/panel_paiements.html', context, request=request
                ),
                'panel-finances': render_to_string(
                    'backoffice/partials/panel_finances.html', context, request=request
                ),
            },
            'charts_json': context['charts_json'],
        }
        return JsonResponse(payload)

    return render(request, 'backoffice/dashboard.html', context)


@backoffice_required
def backoffice_export_excel(request):
    """Export .xlsx des utilisateurs filtrés (mêmes filtres que le tableau, y compris période)."""
    maintenant = timezone.now()
    periode = _periode_depuis_request(request)
    filtre_actif = periode['filtre_actif']
    q, filtre_statut, filtre_plan, non_verifies, user_scope, pay_q, pay_statut, _charge_q, _charge_cat = _filtres_depuis_request(request)
    p_debut, p_fin, scope_effectif = _params_liste_utilisateurs(
        filtre_actif, periode['debut'], periode['fin'], user_scope,
    )
    badge_debut = periode['debut'] if filtre_actif else None
    badge_fin = periode['fin'] if filtre_actif else None
    users_list, emails_verifies, user_ids_verifies = _queryset_utilisateurs_filtres(
        q,
        filtre_statut,
        filtre_plan,
        non_verifies,
        periode_debut=p_debut,
        periode_fin=p_fin,
        user_scope=scope_effectif,
        maintenant=maintenant,
    )
    lignes = _lignes_utilisateurs(
        users_list,
        emails_verifies,
        maintenant,
        user_ids_verifies=user_ids_verifies,
        periode_debut=badge_debut,
        periode_fin=badge_fin,
        filtre_statut=filtre_statut,
        filtre_plan=filtre_plan,
    )

    wb = Workbook()
    ws = wb.active
    ws.title = 'Utilisateurs'
    headers = [
        'E-mail',
        'Nom',
        'Rôle',
        'Organisation',
        'Plan',
        'Statut',
        'Échéance',
        'Téléphone',
        'Dernier paiement',
        'Date paiement',
        'E-mail vérifié',
        'Actif',
        'Dernière connexion',
        'Inscrit le',
        'Inscrit sur période',
        'Connecté sur période',
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for row in lignes:
        org = row['organisation']
        ws.append([
            row['email'],
            row['nom'] or '',
            row['role'] or '',
            (org.nom or org.slug) if org else '',
            row['plan'] if row['plan'] != '—' else '',
            row['statut_label'],
            _date_excel(row['echeance']),
            row['telephone'] or '',
            row['paiement_montant'] or '',
            _dt_excel(row.get('paiement_dt')),
            'Oui' if row['email_verifie'] else 'Non',
            'Oui' if row['actif'] else 'Non',
            _dt_excel(row.get('last_login')),
            _dt_excel(row.get('date_joined')),
            'Oui' if row.get('inscrit_sur_periode') else 'Non',
            'Oui' if row.get('connecte_sur_periode') else 'Non',
        ])

    for col in ws.columns:
        max_len = 0
        letter = col[0].column_letter
        for cell in col:
            val = '' if cell.value is None else str(cell.value)
            max_len = max(max_len, len(val))
        ws.column_dimensions[letter].width = min(max_len + 2, 42)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    stamp = timezone.localtime(maintenant).strftime('%Y%m%d-%H%M')
    filename = f'utilisateurs-xaliss-{stamp}.xlsx'
    response = FileResponse(
        buffer,
        as_attachment=True,
        filename=filename,
        content_type=(
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        ),
    )
    return response


@backoffice_required
def backoffice_export_charges_excel(request):
    """Export .xlsx des charges plateforme (filtres liste + période)."""
    maintenant = timezone.now()
    periode = _periode_depuis_request(request)
    filtre_actif = periode['filtre_actif']
    _q, _fs, _fp, _nv, _us, _pq, _ps, charge_q, charge_categorie = _filtres_depuis_request(request)
    p_debut = periode['debut'] if filtre_actif else None
    p_fin = periode['fin'] if filtre_actif else None
    charges_list = _queryset_charges_filtres(
        charge_q=charge_q,
        charge_categorie=charge_categorie,
        debut=p_debut,
        fin=p_fin,
    )
    cat_labels = dict(ChargePlateforme.CATEGORIE_CHOICES)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Charges Xaliss'
    headers = [
        'Date',
        'Libellé',
        'Catégorie',
        'Montant',
        'Devise',
        'Notes',
        'Créée le',
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for ch in charges_list:
        ws.append([
            ch.date_charge.strftime('%Y-%m-%d'),
            ch.libelle,
            cat_labels.get(ch.categorie, ch.categorie),
            float(ch.montant),
            ch.devise,
            (ch.notes or '').strip(),
            _dt_excel(ch.cree_le),
        ])

    for col in ws.columns:
        max_len = 0
        letter = col[0].column_letter
        for cell in col:
            val = '' if cell.value is None else str(cell.value)
            max_len = max(max_len, len(val))
        ws.column_dimensions[letter].width = min(max_len + 2, 42)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    stamp = timezone.localtime(maintenant).strftime('%Y%m%d-%H%M')
    filename = f'charges-xaliss-{stamp}.xlsx'
    return FileResponse(
        buffer,
        as_attachment=True,
        filename=filename,
        content_type=(
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        ),
    )


def _email_verifie_user(user):
    email = (user.email or '').strip()
    if EmailAddress.objects.filter(user_id=user.pk, verified=True).exists():
        return True
    if email and EmailAddress.objects.filter(
        email__iexact=email, verified=True,
    ).exists():
        return True
    return False


def _detail_paiements_org(org, limite=50):
    qs = (
        PaiementAbonnement.objects
        .filter(organisation=org)
        .select_related('abonnement', 'abonnement__plan')
        .order_by('-cree_le')
    )
    total = qs.count()
    lignes = _lignes_paiements(qs[:limite])
    return {
        'lignes': list(lignes),
        'total': total,
        'limite': limite,
        'tronque': total > limite,
    }


def _detail_membres_org(org, maintenant=None, focus_user_id=None):
    maintenant = maintenant or timezone.now()
    membres = (
        org.membres
        .select_related('utilisateur', 'utilisateur__profil')
        .order_by('role', 'id')
    )
    role_rank = {
        MembreOrganisation.ROLE_PROPRIETAIRE: 0,
        MembreOrganisation.ROLE_ADMIN: 1,
        MembreOrganisation.ROLE_MEMBRE: 2,
    }
    rows = []
    for m in sorted(list(membres), key=lambda x: (role_rank.get(x.role, 9), x.pk)):
        u = m.utilisateur
        rows.append({
            'user_id': u.pk,
            'email': m.get_email(),
            'nom': m.get_nom_affichage() or '',
            'role': m.get_role_display_label(),
            'role_code': m.role,
            'actif': m.actif and u.is_active,
            'is_focus': focus_user_id is not None and u.pk == focus_user_id,
            'login_relatif': _date_relative(u.last_login, maintenant),
            'inscrit_relatif': _date_relative(u.date_joined, maintenant),
            'detail_url': reverse('backoffice_user_detail', args=[u.pk])
            + f'?org={org.pk}',
            'admin_url': f'/admin/auth/user/{u.pk}/change/',
        })
    return rows

def _detail_abonnement_ctx(abo, maintenant=None):
    """Contexte d’affichage abo — lecture pure (pas d’écriture en base)."""
    maintenant = maintenant or timezone.now()
    if not abo:
        return None

    statut_db = abo.statut
    statut_aff = _statut_effectif(abo, maintenant)
    # Pour dates d’affichage d’essai/payant : appliquer les mêmes règles d’exclusivité
    # sans purger la base (les actions POST s’en chargent).
    show_essai, show_periode = _affichage_periodes_abo(abo, maintenant)
    echeance_dt, echeance_label = _echeance_info(abo, maintenant)
    grace = None
    if statut_aff == AbonnementOrganisation.STATUT_EN_RETARD:
        grace = abo.fin_grace_retard()
    plans = list(
        PlanAbonnement.objects.filter(actif=True).order_by('ordre', 'code')
    )
    if not plans:
        plans = list(PlanAbonnement.objects.order_by('ordre', 'code'))
    return {
        'abo': abo,
        'abo_id': abo.pk,
        'plan': abo.plan.nom if abo.plan_id else '—',
        'plan_code': abo.plan.code if abo.plan_id else '',
        'statut': statut_aff,
        'statut_db': statut_db,
        'statut_label': _statut_label(statut_aff),
        'statut_db_label': abo.get_statut_display(),
        'statut_desync': statut_aff != statut_db,
        'prix': _format_montant(abo.plan.prix_mensuel if abo.plan_id else 0),
        'essai_debut': _format_dt_fr(abo.essai_debut),
        'essai_fin': _format_dt_fr(abo.essai_fin),
        'periode_debut': _format_dt_fr(abo.periode_debut),
        'periode_fin': _format_dt_fr(abo.periode_fin),
        'show_essai': show_essai,
        'show_periode': show_periode,
        'echeance': _format_dt_fr(echeance_dt),
        'echeance_label': echeance_label,
        'grace': _format_dt_fr(grace) if grace else '',
        'renouvellement_auto': abo.renouvellement_auto,
        'fournisseur': abo.fournisseur or '',
        'id_externe': abo.id_externe or '',
        'cree_le': _format_dt_fr(abo.cree_le),
        'modifie_le': _format_dt_fr(abo.modifie_le),
        'admin_url': f'/admin/comptes/abonnementorganisation/{abo.pk}/change/',
        'acces_pro': abo.acces_pro_effectif(maintenant),
        'plans_choices': [
            {'code': p.code, 'nom': p.nom, 'selected': abo.plan_id == p.pk}
            for p in plans
        ],
        'statuts_choices': [
            # Select : valeur courante en base (actions forcent un statut DB)
            {'code': code, 'label': label, 'selected': statut_db == code}
            for code, label in AbonnementOrganisation.STATUT_CHOICES
        ],
        'jours_defaut_essai': AbonnementOrganisation.duree_essai().days,
    }


def _detail_org_ctx(org, maintenant=None, focus_user_id=None):
    maintenant = maintenant or timezone.now()
    try:
        abo = org.abonnement
    except AbonnementOrganisation.DoesNotExist:
        abo = None
    membres = _detail_membres_org(org, maintenant, focus_user_id=focus_user_id)
    pay_data = _detail_paiements_org(org)
    stats_pay = _stats_paiements(
        PaiementAbonnement.objects.filter(organisation=org)
    )
    abo_ctx = _detail_abonnement_ctx(abo, maintenant)
    return {
        'org': org,
        'nom': org.nom or org.slug,
        'slug': org.slug,
        'email': org.email or '',
        'telephone': org.telephone or '',
        'adresse': org.adresse or '',
        'site_web': org.site_web or '',
        'devise': org.libelle_devise or 'XOF',
        'cree_le': _format_dt_fr(org.cree_le),
        'modifie_le': _format_dt_fr(org.modifie_le),
        'abonnement': abo_ctx,
        'membres': membres,
        'paiements': pay_data['lignes'],
        'paiements_total': pay_data['total'],
        'paiements_tronque': pay_data['tronque'],
        'paiements_limite': pay_data['limite'],
        'pay_stats': stats_pay,
        'n_membres': len(membres),
        'detail_url': reverse('backoffice_org_detail', args=[org.pk]),
        'admin_url': f'/admin/comptes/organisation/{org.pk}/change/',
    }


def _detail_compte_ctx(user, maintenant=None):
    """Bloc Compte pour la fiche unifiée."""
    maintenant = maintenant or timezone.now()
    try:
        profil = user.profil
    except ProfilUtilisateur.DoesNotExist:
        profil = None
    email = (user.email or user.get_username() or '').strip()
    return {
        'user': user,
        'user_id': user.pk,
        'email': email,
        'nom': user.get_full_name().strip(),
        'username': user.get_username(),
        'actif': user.is_active,
        'staff': user.is_staff,
        'superuser': user.is_superuser,
        'email_verifie': _email_verifie_user(user),
        'en_ligne': user.pk in _ids_users_session_en_cours(),
        'date_joined': _format_dt_fr(user.date_joined),
        'date_joined_relatif': _date_relative(user.date_joined, maintenant),
        'last_login': _format_dt_fr(user.last_login) if user.last_login else 'Jamais',
        'last_login_relatif': _date_relative(user.last_login, maintenant),
        'pays': (profil.pays if profil else '') or '',
        'ville': (profil.ville if profil else '') or '',
        'email_en_attente': (profil.email_en_attente if profil else '') or '',
        'detail_url': reverse('backoffice_user_detail', args=[user.pk]),
        'admin_url': f'/admin/auth/user/{user.pk}/change/',
    }


def _user_orgs_switcher(user, focus_org_id=None):
    """Liste des orgs d’un user pour le sélecteur (une fiche, plusieurs orgs)."""
    membres = (
        user.membres_organisations
        .select_related('organisation')
        .filter(actif=True)
        .order_by('id')
    )
    items = []
    for m in membres:
        org = m.organisation
        items.append({
            'org_id': org.pk,
            'nom': org.nom or org.slug,
            'role': m.get_role_display_label(),
            'active': focus_org_id == org.pk,
            'url': (
                reverse('backoffice_user_detail', args=[user.pk])
                + f'?org={org.pk}'
            ),
        })
    return items


def _build_fiche_unifiee(*, focus, user=None, org=None, org_query=None):
    """
    Une seule page : Compte + Organisation + Abonnement + Membres + Paiements.
    focus: 'user' | 'org'
    """
    maintenant = timezone.now()
    compte = _detail_compte_ctx(user, maintenant) if user else None
    focus_user_id = user.pk if user else None

    org_ctx = None
    orgs_switcher = []
    role_dans_org = ''

    if focus == 'user' and user:
        principal = _membre_principal(user)
        membres_u = list(
            user.membres_organisations
            .select_related(
                'organisation',
                'organisation__abonnement',
                'organisation__abonnement__plan',
            )
            .order_by('-actif', 'id')
        )
        # Choix d’org : ?org= / principale / première
        chosen = None
        if org_query:
            try:
                oid = int(org_query)
            except (TypeError, ValueError):
                oid = None
            if oid:
                for m in membres_u:
                    if m.organisation_id == oid:
                        chosen = m
                        break
        if chosen is None and principal:
            for m in membres_u:
                if m.organisation_id == principal.organisation_id:
                    chosen = m
                    break
        if chosen is None and membres_u:
            chosen = membres_u[0]
        if chosen:
            org_ctx = _detail_org_ctx(
                chosen.organisation,
                maintenant,
                focus_user_id=focus_user_id,
            )
            role_dans_org = chosen.get_role_display_label()
            orgs_switcher = _user_orgs_switcher(user, focus_org_id=chosen.organisation_id)
        fiche_title = compte['email'] or f'#{user.pk}'
        admin_url = compte['admin_url']
        if org_ctx:
            sous_titre = org_ctx['nom']
        else:
            sous_titre = 'Sans organisation'
    else:
        # focus org — un seul chargement org/abo (pas de double GET mutatif)
        owner_m = (
            MembreOrganisation.objects
            .select_related('utilisateur', 'utilisateur__profil')
            .filter(
                organisation=org,
                role=MembreOrganisation.ROLE_PROPRIETAIRE,
                actif=True,
            )
            .first()
        )
        if owner_m is None:
            owner_m = (
                MembreOrganisation.objects
                .select_related('utilisateur', 'utilisateur__profil')
                .filter(organisation=org, actif=True)
                .order_by('id')
                .first()
            )
        focus_user_id = owner_m.utilisateur_id if owner_m else None
        org_ctx = _detail_org_ctx(org, maintenant, focus_user_id=focus_user_id)
        fiche_title = org_ctx['nom']
        admin_url = org_ctx['admin_url']
        if owner_m:
            compte = _detail_compte_ctx(owner_m.utilisateur, maintenant)
            role_dans_org = owner_m.get_role_display_label()
        sous_titre = f"{org_ctx['n_membres']} membre{'s' if org_ctx['n_membres'] != 1 else ''}"

    abo = org_ctx['abonnement'] if org_ctx else None
    if org_ctx:
        org_ctx['abo_action_url'] = reverse(
            'backoffice_org_abo_action', args=[org_ctx['org'].pk]
        )
    if focus == 'user' and user and org_ctx:
        fiche_self_url = (
            reverse('backoffice_user_detail', args=[user.pk])
            + f"?org={org_ctx['org'].pk}"
        )
    elif focus == 'user' and user:
        fiche_self_url = reverse('backoffice_user_detail', args=[user.pk])
    else:
        fiche_self_url = reverse('backoffice_org_detail', args=[org.pk])
    return {
        'fiche_title': fiche_title,
        'focus': focus,
        'focus_label': 'Compte' if focus == 'user' else 'Organisation',
        'sous_titre': sous_titre,
        'compte': compte,
        'org': org_ctx,
        'abo': abo,
        'orgs_switcher': orgs_switcher,
        'role_dans_org': role_dans_org,
        'has_org': org_ctx is not None,
        'admin_url': admin_url,
        'fiche_self_url': fiche_self_url,
        'retour_url': reverse('backoffice') + (
            '#utilisateurs' if focus == 'user' else '#vue'
        ),
        'maintenant': maintenant,
    }


def _safe_bo_next(next_url, fallback):
    nxt = (next_url or '').strip()
    if nxt.startswith('/') and not nxt.startswith('//'):
        return nxt
    return fallback


@backoffice_required
@require_POST
def backoffice_lancement_action(request):
    """
    Jour J : passe toutes les orgs encore en prélaunch en essai Pro (3 mois).
    Ne touche aucune donnée métier (transactions, clients, notes…) — statut abo seulement.
    Les nouvelles inscriptions démarreront ensuite directement en essai
    (via date_lancement_effective).
    """
    fallback = reverse('backoffice') + '#outils'
    next_url = _safe_bo_next(request.POST.get('next'), fallback)
    if '#' not in next_url:
        next_url = f'{next_url}#outils'

    ok = True
    level = 'success'
    msg_text = ''
    updated = 0

    try:
        phrase = (request.POST.get('confirmation') or '').strip().upper()
        if phrase != LANCEMENT_CONFIRM_PHRASE:
            raise ValueError(
                f'Confirmation incorrecte. Tapez exactement « {LANCEMENT_CONFIRM_PHRASE} ».'
            )

        debut = timezone.now()
        qs = (
            AbonnementOrganisation.objects.filter(
                lancement_applique_le__isnull=True,
                statut=AbonnementOrganisation.STATUT_PRELAUNCH,
            )
            .select_related('organisation', 'plan')
            .order_by('pk')
        )
        total = qs.count()
        if total == 0:
            raise ValueError('Aucune organisation en prélaunch à activer.')

        for abo in qs.iterator():
            abo.demarrer_essai(debut=debut, save=True)
            updated += 1

        # Orgs sans abonnement : créer + démarrer l’essai (plateforme déjà lancée).
        for org in Organisation.objects.filter(abonnement__isnull=True).iterator():
            AbonnementOrganisation.creer_pour_organisation(org, save=True)
            updated += 1

        jours = AbonnementOrganisation.duree_essai().days
        msg_text = (
            f'Lancement activé : {updated} organisation(s) passée(s) en essai Pro '
            f'({jours} jours). Aucune donnée métier n’a été modifiée.'
        )
    except ValueError as exc:
        ok = False
        level = 'error'
        msg_text = str(exc)
    except Exception:
        import logging
        logging.getLogger(__name__).exception('backoffice_lancement_action failed')
        ok = False
        level = 'error'
        msg_text = 'Impossible d’activer le lancement. Réessayez.'

    if level == 'success':
        messages.success(request, msg_text)
    else:
        messages.error(request, msg_text)

    if _request_wants_ajax(request):
        return JsonResponse(
            {
                'ok': ok,
                'level': level,
                'message': msg_text,
                'next': next_url,
                'updated': updated,
            }
        )

    return redirect(next_url)


@backoffice_required
@require_POST
def backoffice_broadcast_notif_action(request):
    """Envoie une notification in-app à tous les utilisateurs actifs de Xaliss."""
    from finances.services.notifications import (
        NotificationServiceError,
        broadcast_notification_to_all_users,
    )

    fallback = reverse('backoffice') + '#outils'
    next_url = _safe_bo_next(request.POST.get('next'), fallback)
    if '#' not in next_url:
        next_url = f'{next_url}#outils'

    ok = True
    level = 'success'
    msg_text = ''
    created = 0

    try:
        message = (request.POST.get('message') or '').strip()
        type_notif = (request.POST.get('type') or 'info').strip().lower()
        result = broadcast_notification_to_all_users(
            message=message,
            type_notif=type_notif,
        )
        created = int(result.get('created') or 0)
        if created == 0:
            raise ValueError('Aucun destinataire actif trouvé.')
        msg_text = (
            f'Notification envoyée à {created} utilisateur'
            f'{"s" if created > 1 else ""}.'
        )
    except NotificationServiceError as exc:
        ok = False
        level = 'error'
        msg_text = exc.message
    except ValueError as exc:
        ok = False
        level = 'error'
        msg_text = str(exc)
    except Exception:
        import logging
        logging.getLogger(__name__).exception(
            'backoffice_broadcast_notif_action failed'
        )
        ok = False
        level = 'error'
        msg_text = 'Impossible d’envoyer la notification. Réessayez.'

    if level == 'success':
        messages.success(request, msg_text)
    else:
        messages.error(request, msg_text)

    if _request_wants_ajax(request):
        return JsonResponse(
            {
                'ok': ok,
                'level': level,
                'message': msg_text,
                'next': next_url,
                'created': created,
            }
        )

    return redirect(next_url)


@backoffice_required
@require_POST
def backoffice_sondage_action(request):
    """Crée un sondage à choix unique et le diffuse dans la cloche."""
    from finances.services.notifications import (
        NotificationServiceError,
        create_poll_and_broadcast,
    )

    fallback = reverse('backoffice') + '#outils'
    next_url = _safe_bo_next(request.POST.get('next'), fallback)
    if '#' not in next_url:
        next_url = f'{next_url}#outils'

    ok = True
    level = 'success'
    created = 0
    poll_id = None
    try:
        result = create_poll_and_broadcast(
            question=request.POST.get('question'),
            choices=request.POST.getlist('options'),
            created_by=request.user,
        )
        created = int(result.get('created') or 0)
        poll_id = result.get('poll_id')
        msg_text = (
            f'Sondage envoyé à {created} utilisateur'
            f'{"s" if created > 1 else ""}.'
        )
    except NotificationServiceError as exc:
        ok = False
        level = 'error'
        msg_text = exc.message
    except Exception:
        import logging
        logging.getLogger(__name__).exception('backoffice_sondage_action failed')
        ok = False
        level = 'error'
        msg_text = 'Impossible d’envoyer le sondage. Réessayez.'

    if ok:
        messages.success(request, msg_text)
    else:
        messages.error(request, msg_text)

    if _request_wants_ajax(request):
        return JsonResponse({
            'ok': ok,
            'level': level,
            'message': msg_text,
            'next': next_url,
            'created': created,
            'poll_id': poll_id,
        })
    return redirect(next_url)


@backoffice_required
@require_POST
def backoffice_prolonger_tous_action(request):
    """Prolonge l’accès de toutes les organisations, quel que soit le statut."""
    fallback = reverse('backoffice') + '#outils'
    next_url = _safe_bo_next(request.POST.get('next'), fallback)
    if '#' not in next_url:
        next_url = f'{next_url}#outils'

    ok = True
    level = 'success'
    msg_text = ''
    updated = 0
    jours = 0

    try:
        phrase = (request.POST.get('confirmation') or '').strip().upper()
        if phrase != PROLONGER_TOUS_CONFIRM_PHRASE:
            raise ValueError(
                f'Confirmation incorrecte. Tapez exactement « {PROLONGER_TOUS_CONFIRM_PHRASE} ».'
            )

        jours = AbonnementOrganisation._parse_jours_admin(
            request.POST.get('jours'),
            default=30,
            mini=1,
            maxi=365,
        )
        qs = AbonnementOrganisation.objects.select_related('organisation').order_by('pk')
        total = qs.count()
        if total == 0:
            raise ValueError('Aucun abonnement à prolonger.')

        for abo in qs.iterator():
            abo.prolonger_acces(jours, save=True)
            updated += 1

        msg_text = (
            f'Accès prolongé de {jours} jour{"s" if jours > 1 else ""} '
            f'pour {updated} organisation{"s" if updated > 1 else ""} '
            f'(essai ou période payante selon le statut).'
        )
    except ValueError as exc:
        ok = False
        level = 'error'
        msg_text = str(exc)
    except Exception:
        import logging
        logging.getLogger(__name__).exception(
            'backoffice_prolonger_tous_action failed'
        )
        ok = False
        level = 'error'
        msg_text = 'Impossible de prolonger les accès. Réessayez.'

    if level == 'success':
        messages.success(request, msg_text)
    else:
        messages.error(request, msg_text)

    if _request_wants_ajax(request):
        return JsonResponse(
            {
                'ok': ok,
                'level': level,
                'message': msg_text,
                'next': next_url,
                'updated': updated,
                'jours': jours,
            }
        )

    return redirect(next_url)


@backoffice_required
@require_POST
def backoffice_org_abo_action(request, org_id):
    """
    Actions admin abonnement depuis la fiche unifiée :
    prolonger essai / période, statut, plan, renouvellement, etc.

    Avec X-Requested-With / Accept JSON → JsonResponse (soft reload côté client).
    """
    org = get_object_or_404(Organisation, pk=org_id)
    fallback = reverse('backoffice_org_detail', args=[org.pk]) + '#sec-abonnement'
    next_url = _safe_bo_next(request.POST.get('next'), fallback)
    if '#' not in next_url:
        next_url = f'{next_url}#sec-abonnement'

    action = (request.POST.get('action') or '').strip().lower()
    ok = True
    level = 'success'
    msg_text = ''

    try:
        abo = org.abonnement
    except AbonnementOrganisation.DoesNotExist:
        abo = AbonnementOrganisation.creer_pour_organisation(org, save=True)

    try:
        if action == 'prolonger_essai':
            jours = request.POST.get('jours') or 30
            n = abo.prolonger_essai(jours)
            msg_text = (
                f'Essai prolongé de {n} jour(s). '
                f'Nouvelle fin : {_format_dt_fr(abo.essai_fin)}.'
            )
        elif action == 'redemarrer_essai':
            abo.demarrer_essai()
            msg_text = (
                f'Essai redémarré ({AbonnementOrganisation.duree_essai().days} j). '
                f'Fin : {_format_dt_fr(abo.essai_fin)}.'
            )
        elif action == 'prolonger_periode':
            jours = request.POST.get('jours') or 30
            n = abo.prolonger_periode_payante(jours)
            msg_text = (
                f'Période payante prolongée de {n} jour(s). '
                f'Nouvelle fin : {_format_dt_fr(abo.periode_fin)}.'
            )
        elif action == 'activer_payant':
            jours = request.POST.get('jours') or 30
            n = abo.activer_periode_payante(jours)
            msg_text = (
                f'Période payante activée pour {n} jour(s) '
                f'(jusqu’au {_format_dt_fr(abo.periode_fin)}).'
            )
        elif action == 'definir_statut':
            statut = request.POST.get('statut') or ''
            jours = request.POST.get('jours') or 30
            abo.definir_statut_admin(statut, jours=jours)
            msg_text = f'Statut passé à « {abo.get_statut_display()} ».'
        elif action == 'changer_plan':
            code = request.POST.get('plan') or ''
            abo.changer_plan(code)
            msg_text = f'Plan basculé vers « {abo.plan.nom} ».'
        elif action == 'renouvellement':
            raw = (request.POST.get('value') or '').strip().lower()
            actif = raw in ('1', 'true', 'oui', 'on', 'yes')
            abo.definir_renouvellement_auto(actif)
            msg_text = (
                'Renouvellement auto : '
                + ('activé' if abo.renouvellement_auto else 'désactivé')
                + '.'
            )
        elif action == 'synchroniser':
            abo.normaliser_coherence(save=True, synchroniser=True)
            abo.refresh_from_db()
            msg_text = f'Statut synchronisé → « {abo.get_statut_display()} ».'
        else:
            ok = False
            level = 'error'
            msg_text = 'Action abonnement inconnue.'
    except ValueError as exc:
        ok = False
        level = 'error'
        msg_text = str(exc)
    except Exception:
        import logging
        logging.getLogger(__name__).exception(
            'backoffice_org_abo_action failed org_id=%s action=%s',
            org_id,
            action,
        )
        ok = False
        level = 'error'
        msg_text = (
            'Impossible d’appliquer l’action. Réessayez ou passez par Admin Django.'
        )

    if level == 'success':
        messages.success(request, msg_text)
    else:
        messages.error(request, msg_text)

    if _request_wants_ajax(request):
        return JsonResponse(
            {
                'ok': ok,
                'level': level,
                'message': msg_text,
                'next': next_url,
                'action': action,
            }
        )

    return redirect(next_url)


@backoffice_required
@require_POST
def backoffice_charge_action(request):
    """Créer, modifier ou supprimer une charge plateforme (backoffice ops)."""
    fallback = reverse('backoffice') + '#finances'
    next_url = _safe_bo_next(request.POST.get('next'), fallback)

    action = (request.POST.get('action') or '').strip().lower()
    ok = True
    level = 'success'
    msg_text = ''

    try:
        if action in ('create', 'update'):
            payload = _parse_charge_payload(request.POST)
            if action == 'create':
                ChargePlateforme.objects.create(
                    **payload,
                    cree_par=request.user if request.user.is_authenticated else None,
                )
                msg_text = f'Charge enregistrée : « {payload["libelle"]} ».'
            else:
                charge_id = request.POST.get('charge_id')
                charge = get_object_or_404(ChargePlateforme, pk=charge_id)
                for key, val in payload.items():
                    setattr(charge, key, val)
                charge.save(update_fields=[*payload.keys(), 'modifie_le'])
                msg_text = f'Charge modifiée : « {payload["libelle"]} ».'
        elif action == 'delete':
            charge_id = request.POST.get('charge_id')
            charge = get_object_or_404(ChargePlateforme, pk=charge_id)
            lib = charge.libelle
            charge.delete()
            msg_text = f'Charge supprimée : « {lib} ».'
        else:
            ok = False
            level = 'error'
            msg_text = 'Action charge inconnue.'
    except ValueError as exc:
        ok = False
        level = 'error'
        msg_text = str(exc)
    except Exception:
        import logging
        logging.getLogger(__name__).exception(
            'backoffice_charge_action failed action=%s',
            action,
        )
        ok = False
        level = 'error'
        msg_text = 'Impossible d’appliquer l’action sur la charge.'

    if level == 'success':
        messages.success(request, msg_text)
    else:
        messages.error(request, msg_text)

    if _request_wants_ajax(request):
        return JsonResponse(
            {
                'ok': ok,
                'level': level,
                'message': msg_text,
                'next': next_url,
                'action': action,
            }
        )

    return redirect(next_url)


@backoffice_required
@require_POST
def backoffice_acces_action(request):
    """Ajouter ou révoquer un e-mail autorisé au backoffice (hors .env)."""
    from django.core.exceptions import ValidationError
    from django.core.validators import validate_email

    fallback = reverse('backoffice') + '#outils'
    next_url = _safe_bo_next(request.POST.get('next'), fallback)

    action = (request.POST.get('action') or '').strip().lower()
    email_raw = (request.POST.get('email') or '').strip().lower()
    note = (request.POST.get('note') or '').strip()[:200]
    ok = True
    level = 'success'
    msg_text = ''

    viewer_email = (
        getattr(request.user, 'email', None) or request.user.get_username() or ''
    ).strip().lower()

    try:
        if action == 'add':
            if not email_raw:
                raise ValueError('Indiquez une adresse e-mail.')
            try:
                validate_email(email_raw)
            except ValidationError as exc:
                raise ValueError('Adresse e-mail invalide.') from exc

            row, created = AccesBackoffice.objects.get_or_create(
                email=email_raw,
                defaults={
                    'actif': True,
                    'note': note,
                    'ajoute_par': request.user if request.user.is_authenticated else None,
                },
            )
            if not created:
                row.actif = True
                if note:
                    row.note = note
                if row.ajoute_par_id is None and request.user.is_authenticated:
                    row.ajoute_par = request.user
                row.save()
                msg_text = f'Accès backoffice réactivé pour {email_raw}.'
            else:
                msg_text = f'Accès backoffice accordé à {email_raw}.'

            UserModel = get_user_model()
            if not UserModel.objects.filter(email__iexact=email_raw).exists():
                msg_text += (
                    ' Aucun compte Xaliss avec cet e-mail pour l’instant '
                    '— l’accès s’appliquera dès la première connexion.'
                )
            invalider_cache_acces_backoffice()

        elif action == 'revoke':
            if not email_raw:
                raise ValueError('E-mail manquant.')
            if email_raw == viewer_email:
                raise ValueError('Vous ne pouvez pas révoquer votre propre accès.')

            row = AccesBackoffice.objects.filter(email__iexact=email_raw).first()
            if not row or not row.actif:
                raise ValueError('Cet accès n’est pas actif dans le backoffice.')

            row.actif = False
            row.save(update_fields=['actif', 'modifie_le'])
            invalider_cache_acces_backoffice()

            if email_raw in emails_backoffice_env():
                msg_text = (
                    f'Accès base révoqué pour {email_raw}, mais l’e-mail reste '
                    f'autorisé via BACKOFFICE_ALLOWED_EMAILS (.env).'
                )
            else:
                msg_text = f'Accès backoffice révoqué pour {email_raw}.'

        elif action == 'reactivate':
            if not email_raw:
                raise ValueError('E-mail manquant.')
            row = AccesBackoffice.objects.filter(email__iexact=email_raw).first()
            if not row:
                raise ValueError('Aucun enregistrement à réactiver.')
            row.actif = True
            row.save(update_fields=['actif', 'modifie_le'])
            invalider_cache_acces_backoffice()
            msg_text = f'Accès backoffice réactivé pour {email_raw}.'

        elif action == 'delete':
            if not email_raw:
                raise ValueError('E-mail manquant.')
            if email_raw == viewer_email:
                raise ValueError('Vous ne pouvez pas supprimer votre propre accès.')

            row = AccesBackoffice.objects.filter(email__iexact=email_raw).first()
            if not row:
                raise ValueError('Aucun enregistrement backoffice à supprimer.')

            row.delete()
            invalider_cache_acces_backoffice()

            if email_raw in emails_backoffice_env():
                msg_text = (
                    f'Entrée supprimée pour {email_raw}. '
                    f'L’e-mail reste autorisé via le .env.'
                )
            else:
                msg_text = f'Accès backoffice supprimé définitivement : {email_raw}.'
        else:
            ok = False
            level = 'error'
            msg_text = 'Action accès inconnue.'
    except ValueError as exc:
        ok = False
        level = 'error'
        msg_text = str(exc)
    except Exception:
        import logging
        logging.getLogger(__name__).exception(
            'backoffice_acces_action failed action=%s',
            action,
        )
        ok = False
        level = 'error'
        msg_text = 'Impossible de modifier les accès. Réessayez.'

    if level == 'success':
        messages.success(request, msg_text)
    else:
        messages.error(request, msg_text)

    if _request_wants_ajax(request):
        return JsonResponse(
            {
                'ok': ok,
                'level': level,
                'message': msg_text,
                'next': next_url,
                'action': action,
            }
        )

    return redirect(next_url)


@backoffice_required
def backoffice_user_detail(request, user_id):
    """Fiche unifiée ouverte depuis un compte."""
    user = get_object_or_404(
        User.objects.select_related('profil'),
        pk=user_id,
    )
    context = _build_fiche_unifiee(
        focus='user',
        user=user,
        org_query=(request.GET.get('org') or '').strip(),
    )
    return render(request, 'backoffice/detail.html', context)


@backoffice_required
def backoffice_org_detail(request, org_id):
    """Fiche unifiée ouverte depuis une organisation."""
    org = get_object_or_404(Organisation, pk=org_id)
    context = _build_fiche_unifiee(focus='org', org=org)
    return render(request, 'backoffice/detail.html', context)